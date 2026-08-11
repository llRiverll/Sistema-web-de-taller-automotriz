# ==============================================================================
# 1. IMPORTACIONES (Ordenadas y limpias)
# ==============================================================================
import os
import random
import smtplib
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
from functools import wraps
from email.message import EmailMessage
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from flask import request, make_response
from werkzeug.utils import secure_filename
from flask import Response

# Flask y utilidades web
from flask import Flask, render_template, request, redirect, url_for, flash, session, make_response, jsonify
from werkzeug.security import generate_password_hash, check_password_hash
from flask import request
# Base de datos
import mysql.connector
from mysql.connector import Error

# Librerías de terceros (PDF, Códigos QR, IA, Letras)
from fpdf import FPDF
import qrcode
from num2words import num2words
import requests
import google.generativeai as genai
import csv
import io
import tempfile
import time
import hashlib
import uuid
# ==============================================================================
# 2. CONFIGURACIÓN PRINCIPAL DE LA APLICACIÓN
# ==============================================================================
app = Flask(__name__)
app.secret_key = 'TallerVirgenDeChapi_Tesis2024_SuperSegura'

# Configuración de la Inteligencia Artificial (Gemini)
# Nota para tu tesis: En un entorno real, esta API Key debería ir en un archivo .env oculto
genai.configure(api_key="AIzaSyD26oW4uirIBAMOSdB7wvJQrtOTeEno4HI")

# --- DATOS DE LA EMPRESA ---
EMPRESA_DATOS = {
    "nombre": "Multiservicios Virgen de Chapi",
    "ruc": "10806325631",
    "direccion": "Carretera a Puquio KM. 2.5, Pj. Buena Fe",
    "ciudad": "Nasca - Ica - Peru",
    "telefono": "966633037 / 056-523915",
    "email": "moralescaritas_@hotmail.com",
    "propietario": "Félix Morales Caritas",
    "bcp": "440-91611954-0-49",
    "serie_boleta": "B001",
    "serie_factura": "F001"
}

# Configuración de la base de datos
DB_CONFIG = {
    'host': 'localhost',
    'user': 'root',
    'password': '',
    'database': 'taller_db'
}

# ==============================================================================
# CONFIGURACIÓN DEL FACTURADOR ELECTRÓNICO (API PSE)
# ==============================================================================
# Nota: Reemplazar con las credenciales reales de producción o pruebas de tu PSE
API_FACTURACION_URL = "https://api.nubefact.com/api/v1/fe/generar"
API_FACTURACION_TOKEN = "TU_TOKEN_DE_SEGURIDAD_AQUI"

# ==============================================================================
# 3. FUNCIONES CORE Y DE SEGURIDAD
# ==============================================================================

# Función para conectar a la Base de Datos
def get_db_connection():
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        if conn.is_connected(): 
            return conn
    except Error as e:
        print(f"Error conectando a BD Taller: {e}")
        return None

# Decorador 1: Exigir que el usuario haya iniciado sesión (El que te faltaba)
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Comprueba si existe 'id' o 'user_id' en la sesión actual
        if 'id' not in session and 'user_id' not in session:
            flash("Por favor, inicie sesión para acceder a esta página.", "warning")
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# Decorador 2: Exigir roles específicos para entrar a un módulo
def roles_required(roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            # 1. Si no hay NADIE logueado, lo pateamos al Login
            if 'user_id' not in session and 'id' not in session:
                return redirect(url_for('login'))
                
            # 2. Si sí está logueado, obtenemos su rol
            user_rol = session.get('rol') or session.get('user_rol')
            
            # 3. Verificamos si su rol está en la lista de permitidos
            if user_rol not in roles:
                flash("No tienes los permisos de sistema necesarios para ver esta sección.", "danger")
                return redirect(url_for('dashboard')) 
            return f(*args, **kwargs)
        return decorated_function
    return decorator

# ==============================================================================
# 1. FUNCIÓN ESPÍA: REGISTRO DE AUDITORÍA (NIVEL BANCARIO)
# ==============================================================================
def registrar_auditoria(accion, modulo, detalles=None):
    usuario_id = session.get('user_id') or session.get('id')
    if not usuario_id:
        usuario_id = None 

    direccion_ip = request.environ.get('HTTP_X_FORWARDED_FOR', request.remote_addr)
    detalles_json = json.dumps(detalles) if detalles else None

    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO auditoria_logs (usuario_id, accion, modulo, direccion_ip, detalles)
                VALUES (%s, %s, %s, %s, %s)
            """, (usuario_id, accion, modulo, direccion_ip, detalles_json))
            conn.commit()
        except Error as e:
            print(f"Error interno guardando Log: {e}")
        finally:
            cursor.close()
            conn.close()

# ==============================================================================
# 2. PANTALLA PRINCIPAL: VISOR DE AUDITORÍA
# ==============================================================================
@app.route('/auditoria')
@roles_required(['admin'])
def ver_auditoria():
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a BD.", "danger")
        return redirect(url_for('dashboard'))

    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT a.id, a.accion, a.modulo, a.direccion_ip, a.fecha_hora, a.detalles,
               p.nombre_completo, u.rol
        FROM auditoria_logs a
        LEFT JOIN usuarios u ON a.usuario_id = u.id
        LEFT JOIN personas p ON u.persona_id = p.id
        ORDER BY a.fecha_hora DESC
        LIMIT 500
    """)
    logs = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('usuarios/auditoria.html', logs=logs)

# ==============================================================================
# 3. REPORTE FORENSE: EXCEL REAL Y FORMATEADO (.XLSX)
# ==============================================================================
@app.route('/auditoria/exportar')
@roles_required(['admin'])
def exportar_auditoria():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT a.fecha_hora, p.nombre_completo, u.rol, a.modulo, a.accion, a.direccion_ip, a.detalles
        FROM auditoria_logs a
        LEFT JOIN usuarios u ON a.usuario_id = u.id
        LEFT JOIN personas p ON u.persona_id = p.id
        ORDER BY a.fecha_hora DESC
    """)
    logs = cursor.fetchall()
    cursor.close()
    conn.close()

    # CREACIÓN DEL EXCEL PROFESIONAL
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría Forense"

    # Estilos Premium para Excel
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")

    # Título Principal
    ws.merge_cells('A1:G1')
    ws['A1'] = "REPORTE FORENSE DE AUDITORÍA - VIRGEN DE CHAPI"
    ws['A1'].font = Font(size=14, bold=True, color="0F172A")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    # Cabeceras
    headers = ['Fecha y Hora', 'Usuario Responsable', 'Rol', 'Módulo', 'Acción Realizada', 'Dirección IP', 'Detalles (JSON)']
    ws.append([]) # Fila vacía para espacio
    ws.append(headers)

    # Pintar Cabeceras
    for col_num, cell in enumerate(ws[3], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border

    # Insertar Datos
    for log in logs:
        fecha = log['fecha_hora'].strftime('%Y-%m-%d %H:%M:%S') if log['fecha_hora'] else 'Desconocida'
        usuario = log['nombre_completo'] or 'Sistema'
        rol = str(log['rol'] or 'N/A').upper()
        modulo = log['modulo']
        accion = log['accion']
        ip = log['direccion_ip'] or 'Localhost'
        detalles = log['detalles'] or '---'

        row = [fecha, usuario, rol, modulo, accion, ip, detalles]
        ws.append(row)
        
        # Poner borde a la fila
        for cell in ws[ws.max_row]:
            cell.border = border

    # Ensanchar columnas automáticamente
    column_widths = {'A': 20, 'B': 35, 'C': 15, 'D': 20, 'E': 45, 'F': 18, 'G': 60}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Convertir a bytes para descargar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Reporte_Auditoria.xlsx"}
    )

# ==============================================================================
# 4. INTELIGENCIA ARTIFICIAL: RESUMEN GERENCIAL (CAMBIO A GROQ / LLaMA 3)
# ==============================================================================
@app.route('/auditoria/resumen_ia')
@roles_required(['admin'])
def resumen_ia_auditoria():
    try:
        # 1. Traemos los últimos 30 movimientos
        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT a.fecha_hora, a.accion, a.modulo, p.nombre_completo
            FROM auditoria_logs a
            LEFT JOIN usuarios u ON a.usuario_id = u.id
            LEFT JOIN personas p ON u.persona_id = p.id
            ORDER BY a.fecha_hora DESC LIMIT 30
        """)
        logs = cursor.fetchall()
        cursor.close()
        conn.close()

        if not logs:
            return jsonify({"status": "error", "message": "No hay registros suficientes para analizar."})

        # 2. Traducimos los datos a texto
        texto_logs = ""
        for l in logs:
            fecha = l['fecha_hora'].strftime('%Y-%m-%d %H:%M') if l['fecha_hora'] else 'Desconocida'
            usuario = l['nombre_completo'] or 'Sistema'
            texto_logs += f"- {fecha} | Usuario: {usuario} | Módulo: {l['modulo']} | Acción: {l['accion']}\n"

        # 3. El Prompt (Instrucciones de Alto Nivel para la IA)
        prompt = f"""
        Eres el Auditor Forense Principal de Multiservicios Virgen de Chapi.
        Analiza los siguientes registros y genera un Resumen Ejecutivo.
        
        REGLA ESTRICTA: Tu respuesta debe estar formateada directamente en HTML (usa <b>, <br>, <ul>, <li>). NO uses Markdown ni asteriscos.
        
        Estructura el informe en 3 partes exactas:
        <b>1. Resumen Operativo:</b> Un párrafo breve de cómo estuvo el día.
        <b>2. Movimientos Clave:</b> Usa una lista con viñetas (<ul><li>) para resumir los cobros, clientes nuevos o reparaciones terminadas.
        <b>3. Análisis de Riesgo:</b> Si hay descuadres, cambios de IGV o eliminaciones, destácalo usando <span style='color:red;'><b>texto en rojo</b></span>. Si el día fue normal, indícalo claramente con un mensaje de tranquilidad.

        Registros recientes:
        {texto_logs}
        """

        # 4. CONEXIÓN DIRECTA CON GROQ (META LLaMA 3)
        import requests
        
        # ⚠️ PON AQUÍ TU API KEY DE GROQ (Empieza con gsk_) ⚠️
        GROQ_API_KEY = "TU_CLAVE_API_DE_GROQ_AQUI"

        url = "https://api.groq.com/openai/v1/chat/completions"
        
        headers = {
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json"
        }
        
        # Usamos LLaMA 3 de 70 billones de parámetros (Súper Inteligente y rápido)
        payload = {
            "model": "llama-3.3-70b-versatile",
            "messages": [
                {"role": "system", "content": "Eres el Gerente Auditor experto de un Taller Automotriz en Perú. Respondes siempre en español."},
                {"role": "user", "content": prompt}
            ],
            "temperature": 0.5
        }
        
        respuesta = requests.post(url, headers=headers, json=payload)
        
        if respuesta.status_code == 200:
            datos = respuesta.json()
            texto_final = datos['choices'][0]['message']['content']
            return jsonify({"status": "success", "resumen": texto_final})
        else:
            return jsonify({"status": "error", "message": f"Error de Groq: {respuesta.text}"})

    except Exception as e:
        return jsonify({"status": "error", "message": f"Error interno: {str(e)}"})

# ==============================================================================
# CONFIGURACIÓN GLOBAL DE LA EMPRESA (PANEL MAESTRO)
# ==============================================================================
@app.route('/configuracion', methods=['GET', 'POST'])
@roles_required(['admin'])
def configuracion_empresa():
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos.", "danger")
        return redirect(url_for('dashboard'))

    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        try:
            # 1. Carga del Logotipo
            if 'logo_empresa' in request.files:
                file = request.files['logo_empresa']
                if file and file.filename != '':
                    ruta_logo = os.path.join('static', 'logo_chapi.png')
                    file.save(ruta_logo)

            # 2. Captura de Datos Generales y Bancarios
            nombre = request.form.get('nombre_taller', '').strip()
            ruc = request.form.get('ruc', '').strip()
            direccion = request.form.get('direccion', '').strip()
            telefono = request.form.get('telefono', '').strip()
            correo = request.form.get('correo', '').strip()
            cuenta_banco = request.form.get('cuenta_banco', '').strip()
            cuenta_yape = request.form.get('cuenta_yape', '').strip()
            mensaje = request.form.get('mensaje_ticket', '').strip()

            # 3. Facturación Electrónica y PSE
            serie_boleta = request.form.get('serie_boleta', 'B001').strip().upper()
            serie_factura = request.form.get('serie_factura', 'F001').strip().upper()
            igv = float(request.form.get('porcentaje_igv', 18.00))
            moneda = request.form.get('moneda', 'S/').strip()
            pse_url = request.form.get('pse_url', '').strip()
            pse_token = request.form.get('pse_token', '').strip()

            # 4. Servidor de Correos SMTP
            smtp_email = request.form.get('smtp_email', '').strip()
            smtp_password = request.form.get('smtp_password', '').strip()

            # 5. Seguridad y Permisos (Checks - Vienen como 'on' si están marcados)
            permitir_credito = 1 if request.form.get('permitir_credito') else 0
            cajeros_pueden_anular = 1 if request.form.get('cajeros_pueden_anular') else 0
            obligar_2fa = 1 if request.form.get('obligar_2fa') else 0
            envio_correos_auto = 1 if request.form.get('envio_correos_auto') else 0

            # Actualización en BD
            cursor.execute("""
                UPDATE configuracion 
                SET nombre_taller = %s, ruc = %s, direccion = %s, 
                    telefono = %s, correo = %s, cuenta_banco = %s, cuenta_yape = %s, mensaje_ticket = %s,
                    serie_boleta = %s, serie_factura = %s, porcentaje_igv = %s, moneda = %s,
                    pse_url = %s, pse_token = %s, smtp_email = %s, smtp_password = %s,
                    permitir_credito = %s, cajeros_pueden_anular = %s, obligar_2fa = %s, envio_correos_auto = %s
                WHERE id = 1
            """, (nombre, ruc, direccion, telefono, correo, cuenta_banco, cuenta_yape, mensaje,
                  serie_boleta, serie_factura, igv, moneda, pse_url, pse_token, smtp_email, smtp_password,
                  permitir_credito, cajeros_pueden_anular, obligar_2fa, envio_correos_auto))
            
            conn.commit()
            
            # Auditoría arreglada
            registrar_auditoria("Actualizó configuración global del sistema", "ADMINISTRACIÓN", {"ajustes": "Se actualizaron parámetros maestros y seguridad"})
            
            flash("✅ Configuración y permisos actualizados con éxito en el sistema.", "success")
            
        except Error as e:
            conn.rollback()
            flash(f"Error al actualizar la configuración: {e.msg}", "danger")
        except ValueError:
            flash("El valor del IGV debe ser numérico.", "danger")

    # GET: Cargar los datos actuales
    cursor.execute("SELECT * FROM configuracion WHERE id = 1")
    config = cursor.fetchone()
    cursor.close()
    conn.close()

    return render_template('configuracion.html', config=config)

# ==============================================================================
# 4. RUTAS DE LA APLICACIÓN (Aquí empiezan tus @app.route...)
# ==============================================================================

@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        # =====================================================================
        # 1. VALIDACIÓN DE BLOQUEO DEL SISTEMA
        # =====================================================================
        if 'bloqueo_hasta' in session:
            # Revisa si la hora actual es menor a la hora en que se levantará el castigo
            if datetime.now() < session['bloqueo_hasta']:
                # Calcula cuántos minutos le faltan de castigo
                tiempo_restante = (session['bloqueo_hasta'] - datetime.now()).seconds // 60
                flash(f"🚫 Sistema bloqueado por seguridad. Intente nuevamente en {tiempo_restante + 1} minuto(s).", "danger")
                return render_template('login.html')
            else:
                # Ya pasó el tiempo de castigo, limpiamos la memoria
                session.pop('bloqueo_hasta', None)
                session.pop('intentos_fallidos', None)
        # =====================================================================

        usuario = request.form['usuario']
        password_ingresada = request.form['password']
        
        # AUDITORÍA: Capturamos IP y Navegador
        ip_cliente = request.headers.get('X-Forwarded-For', request.remote_addr)
        navegador = request.headers.get('User-Agent', '')[:250]
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT u.id, u.usuario, u.password, u.rol, u.activo, 
                       p.nombre_completo as nombre, p.email 
                FROM usuarios u
                JOIN personas p ON u.persona_id = p.id
                WHERE u.usuario = %s AND u.activo = 1
            """, (usuario,))
            user = cursor.fetchone()
            
            if user:
                login_exitoso = False
                password_bd = user['password']

                # LÓGICA DE ENCRIPTACIÓN
                if password_bd.startswith('pbkdf2:sha256:') or password_bd.startswith('scrypt:'):
                    if check_password_hash(password_bd, password_ingresada):
                        login_exitoso = True
                else:
                    if password_bd == password_ingresada:
                        login_exitoso = True
                        nuevo_hash = generate_password_hash(password_ingresada)
                        cursor.execute("UPDATE usuarios SET password = %s WHERE id = %s", (nuevo_hash, user['id']))
                        conn.commit()

                if login_exitoso:
                    # ÉXITO: Limpiamos cualquier error previo
                    session.pop('intentos_fallidos', None)
                    session.pop('bloqueo_hasta', None)
                    
                    cursor.execute("INSERT INTO logs_acceso (usuario, ip_address, exito, navegador) VALUES (%s, %s, 1, %s)", (usuario, ip_cliente, navegador))
                    conn.commit()

                    # VERIFICAR COOKIE
                    nombre_cookie = f"dispositivo_confiable_{user['id']}"
                    if request.cookies.get(nombre_cookie) == 'true':
                        session['user_id'] = user['id']
                        session['user_nombre'] = user['nombre']
                        session['user_rol'] = user['rol']
                        conn.close()
                        flash(f"¡Bienvenido de nuevo, {user['nombre']}!", "success")
                        return redirect(url_for('dashboard'))

                    # Generar Token 2FA
                    token_2fa = str(random.randint(100000, 999999))
                    expiracion = datetime.now() + timedelta(minutes=5) 
                    
                    cursor.execute("UPDATE usuarios SET otp_code = %s, otp_expiry = %s WHERE id = %s", 
                                   (token_2fa, expiracion, user['id']))
                    conn.commit()
                    session['pending_user_id'] = user['id']
                    
                    # ENVÍO DE CORREO 2FA
                    correo_destino = user['email']
                    correo_remitente = 'arturo20010803@gmail.com'
                    password_app = 'wxvqbwoflnoojhqh' 
                    
                    if correo_destino: 
                        try:
                            msg = EmailMessage()
                            msg.set_content(f"Hola {user['nombre']},\n\nTu código de acceso (2FA) es: {token_2fa}\n\nDetectamos un acceso desde la IP: {ip_cliente}.\nEste código expira en 5 minutos.")
                            msg['Subject'] = 'Código de Seguridad 2FA - Virgen de Chapi'
                            msg['From'] = correo_remitente
                            msg['To'] = correo_destino
                            
                            server = smtplib.SMTP('smtp.gmail.com', 587)
                            server.starttls()
                            server.login(correo_remitente, password_app)
                            server.send_message(msg)
                            server.quit()
                        except Exception as e:
                            flash("Error al enviar el correo.", "danger")
                    
                    conn.close()
                    return redirect(url_for('verificar_2fa'))
                
                else:
                    # =====================================================================
                    # 2. LÓGICA DE INTENTOS FALLIDOS (CONTADOR)
                    # =====================================================================
                    cursor.execute("INSERT INTO logs_acceso (usuario, ip_address, exito, navegador) VALUES (%s, %s, 0, %s)", (usuario, ip_cliente, navegador))
                    conn.commit()
                    
                    intentos = session.get('intentos_fallidos', 0) + 1
                    session['intentos_fallidos'] = intentos
                    
                    if intentos >= 3:
                        # Castigamos al usuario bloqueándolo por 5 minutos
                        session['bloqueo_hasta'] = datetime.now() + timedelta(minutes=5)
                        flash("🚫 Ha superado el límite de intentos. El sistema ha sido bloqueado por 5 minutos.", "danger")
                    else:
                        flash(f"❌ Contraseña incorrecta. Le quedan {3 - intentos} intento(s).", "danger")
            else:
                # Log de Fallo (Usuario no existe)
                cursor.execute("INSERT INTO logs_acceso (usuario, ip_address, exito, navegador) VALUES (%s, %s, 0, %s)", (usuario, ip_cliente, navegador))
                conn.commit()
                flash("❌ El usuario no existe o está inactivo.", "danger")
                
            conn.close()
            
    return render_template('login.html') 

@app.route('/verificar-2fa', methods=['GET', 'POST'])
def verificar_2fa():
    if 'pending_user_id' not in session:
        return redirect(url_for('login'))
        
    if request.method == 'POST':
        codigo_ingresado = request.form['codigo_2fa']
        recordar_dispositivo = request.form.get('recordar_dispositivo')
        pending_id = session['pending_user_id']
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("""
                SELECT u.id, u.otp_code, u.otp_expiry, u.rol, p.nombre_completo as nombre
                FROM usuarios u
                JOIN personas p ON u.persona_id = p.id
                WHERE u.id = %s
            """, (pending_id,))
            user = cursor.fetchone()
            
            if user:
                if user['otp_code'] == codigo_ingresado:
                    if datetime.now() <= user['otp_expiry']:
                        cursor.execute("UPDATE usuarios SET otp_code = NULL, otp_expiry = NULL WHERE id = %s", (pending_id,))
                        conn.commit()
                        
                        session.pop('pending_user_id', None)
                        session['user_id'] = user['id']
                        session['user_nombre'] = user['nombre']
                        session['user_rol'] = user['rol']
                        conn.close()
                        
                        flash(f"¡Bienvenido, {user['nombre']}!", "success")
                        
                        respuesta = make_response(redirect(url_for('dashboard')))
                        if recordar_dispositivo:
                            nombre_cookie = f"dispositivo_confiable_{user['id']}"
                            respuesta.set_cookie(nombre_cookie, 'true', max_age=2592000, httponly=True)
                        return respuesta
                    else:
                        flash("El código token ha expirado.", "danger")
                        session.pop('pending_user_id', None)
                        return redirect(url_for('login'))
                else:
                    flash("El código token es incorrecto.", "danger")
            conn.close()
            
    return render_template('verificar_2fa.html')

# --- RUTA PARA CERRAR SESIÓN ---
@app.route('/logout')
def logout():
    session.clear() 
    # Opcional: NO borramos la cookie aquí para que el dispositivo siga siendo de confianza
    flash("Has cerrado sesión de forma segura.", "info")
    return redirect(url_for('login'))

# ================== MÓDULO DE RECUPERACIÓN DE CONTRASEÑA ==================

@app.route('/recuperar-password', methods=['GET', 'POST'])
def recuperar_password():
    if request.method == 'POST':
        email_ingresado = request.form['email'].strip().lower()
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            # Buscamos si existe un usuario activo con ese correo
            cursor.execute("""
                SELECT u.id, p.nombre_completo as nombre, p.email 
                FROM usuarios u
                JOIN personas p ON u.persona_id = p.id
                WHERE p.email = %s AND u.activo = 1
            """, (email_ingresado,))
            user = cursor.fetchone()
            
            if user:
                # 1. Generamos token de rescate de 6 dígitos
                token_rescate = str(random.randint(100000, 999999))
                expiracion = datetime.now() + timedelta(minutes=15) # Validez: 15 min
                
                # 2. Guardamos en BD
                cursor.execute("UPDATE usuarios SET otp_code = %s, otp_expiry = %s WHERE id = %s", 
                               (token_rescate, expiracion, user['id']))
                conn.commit()
                
                session['reset_user_id'] = user['id']
                
                # 3. Enviamos Correo
                correo_remitente = 'arturo20010803@gmail.com'
                password_app = 'wxvqbwoflnoojhqh'
                
                try:
                    msg = EmailMessage()
                    msg.set_content(f"Hola {user['nombre']},\n\nHas solicitado restablecer tu contraseña en el sistema Virgen de Chapi.\n\nTu código de seguridad es: {token_rescate}\n\nIngresa este código en el sistema para crear una nueva clave. Caduca en 15 minutos.")
                    msg['Subject'] = 'Recuperación de Contraseña - Virgen de Chapi'
                    msg['From'] = correo_remitente
                    msg['To'] = email_ingresado
                    
                    server = smtplib.SMTP('smtp.gmail.com', 587)
                    server.starttls()
                    server.login(correo_remitente, password_app)
                    server.send_message(msg)
                    server.quit()
                    
                    flash("Te hemos enviado un código de recuperación a tu correo.", "info")
                    conn.close()
                    return redirect(url_for('reset_password'))
                except Exception as e:
                    print(f"Error correo: {e}")
                    flash("Error al enviar el correo de recuperación.", "danger")
            else:
                # Anti-Hacking: No revelamos si el correo existe o no
                flash("Si el correo existe en nuestro sistema, te enviaremos las instrucciones.", "info")
            
            conn.close()
            
    return render_template('recuperar_password.html')

@app.route('/reset-password', methods=['GET', 'POST'])
def reset_password():
    if 'reset_user_id' not in session:
        return redirect(url_for('login'))
        
    if request.method == 'POST':
        codigo_ingresado = request.form['codigo_rescate']
        nueva_password = request.form['nueva_password']
        user_id = session['reset_user_id']
        
        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            cursor.execute("SELECT otp_code, otp_expiry FROM usuarios WHERE id = %s", (user_id,))
            user = cursor.fetchone()
            
            if user and user['otp_code'] == codigo_ingresado:
                if datetime.now() <= user['otp_expiry']:
                    # ÉXITO: Encriptamos la nueva clave
                    nueva_clave_hash = generate_password_hash(nueva_password)
                    
                    cursor.execute("""
                        UPDATE usuarios 
                        SET password = %s, otp_code = NULL, otp_expiry = NULL 
                        WHERE id = %s
                    """, (nueva_clave_hash, user_id))
                    conn.commit()
                    
                    session.pop('reset_user_id', None)
                    flash("✅ Contraseña actualizada. Ya puedes iniciar sesión.", "success")
                    conn.close()
                    return redirect(url_for('login'))
                else:
                    flash("❌ El código de recuperación ha expirado.", "danger")
            else:
                flash("❌ El código ingresado es incorrecto.", "danger")
                
            conn.close()
            
    return render_template('reset_password.html')

@app.route('/reenviar-codigo', methods=['POST'])
def reenviar_codigo():
    if 'reset_user_id' not in session:
        flash("Tu sesión ha expirado. Vuelve a ingresar tu correo.", "danger")
        return redirect(url_for('recuperar_password'))
        
    user_id = session['reset_user_id']
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT p.nombre_completo as nombre, p.email 
            FROM usuarios u
            JOIN personas p ON u.persona_id = p.id
            WHERE u.id = %s
        """, (user_id,))
        user = cursor.fetchone()
        
        if user:
            # Generamos un NUEVO código
            token_rescate = str(random.randint(100000, 999999))
            expiracion = datetime.now() + timedelta(minutes=15)
            
            cursor.execute("UPDATE usuarios SET otp_code = %s, otp_expiry = %s WHERE id = %s", 
                           (token_rescate, expiracion, user_id))
            conn.commit()
            
            # Reenviamos el correo
            correo_remitente = 'arturo20010803@gmail.com'
            password_app = 'wxvqbwoflnoojhqh'
            try:
                msg = EmailMessage()
                msg.set_content(f"Hola {user['nombre']},\n\nAquí tienes tu NUEVO código de seguridad: {token_rescate}\n\nEste código caducará en 15 minutos.")
                msg['Subject'] = 'NUEVO Código de Seguridad - Virgen de Chapi'
                msg['From'] = correo_remitente
                msg['To'] = user['email']
                
                server = smtplib.SMTP('smtp.gmail.com', 587)
                server.starttls()
                server.login(correo_remitente, password_app)
                server.send_message(msg)
                server.quit()
                flash("✅ Nuevo código enviado exitosamente a tu correo.", "success")
            except Exception as e:
                flash("❌ Error al reenviar el correo. Revisa tu conexión.", "danger")
        conn.close()
        
    return redirect(url_for('reset_password'))


# ================== RUTAS DEL SISTEMA (AHORA PROTEGIDAS) ==================

# --- RUTA PRINCIPAL (Todos pueden verla) ---
@app.route('/dashboard')
@roles_required(['admin', 'mecanico', 'cajero'])
def dashboard():
    conn = get_db_connection()
    kpis = {'autos_taller': 0, 'entregas_hoy': 0, 'ventas_hoy': 0.0, 'alertas_stock': 0}
    chart_labels = []
    chart_data = []
    
    # Nuevas listas para el Centro de Notificaciones
    alertas_repuestos = [] 
    alertas_retrasos = []
    alertas_inventario = []

    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            
            # --- 1. KPIs BÁSICOS ---
            cursor.execute("SELECT COUNT(*) as total FROM ordenes_servicio WHERE estado IN ('pendiente', 'en_proceso', 'esperando_repuesto')")
            kpis['autos_taller'] = cursor.fetchone()['total']
            
            cursor.execute("SELECT COUNT(*) as total FROM ordenes_servicio WHERE DATE(fecha_promesa) = CURDATE() AND estado != 'entregado'")
            kpis['entregas_hoy'] = cursor.fetchone()['total']
            
            cursor.execute("SELECT SUM(monto_total) as total FROM pagos WHERE DATE(fecha_pago) = CURDATE() AND estado != 'anulado'")
            resultado_ventas = cursor.fetchone()
            kpis['ventas_hoy'] = resultado_ventas['total'] if resultado_ventas['total'] else 0.00
            
            if session.get('user_rol') == 'admin':
                cursor.execute("SELECT COUNT(*) as total FROM productos WHERE tipo = 'repuesto' AND stock_actual <= stock_minimo")
                kpis['alertas_stock'] = cursor.fetchone()['total']
            
            # --- 2. ALERTAS DEL PANEL DE CONTROL ---
            
            # A. Vehículos esperando repuestos
            if session.get('user_rol') in ['admin', 'mecanico']:
                cursor.execute("""
                    SELECT o.id, o.repuesto_faltante, o.fecha_pausa, v.placa 
                    FROM ordenes_servicio o
                    JOIN vehiculos v ON o.vehiculo_id = v.id
                    WHERE o.estado = 'esperando_repuesto'
                """)
                alertas_repuestos = cursor.fetchall()

            # B. Vehículos con retraso (Fecha promesa vencida)
            if session.get('user_rol') in ['admin', 'mecanico']:
                cursor.execute("""
                    SELECT o.id, v.placa, o.fecha_promesa, p.nombre_completo as cliente
                    FROM ordenes_servicio o
                    JOIN vehiculos v ON o.vehiculo_id = v.id
                    JOIN clientes c ON v.cliente_id = c.id
                    JOIN personas p ON c.persona_id = p.id
                    WHERE o.estado IN ('pendiente', 'en_proceso') 
                    AND o.fecha_promesa < NOW()
                    ORDER BY o.fecha_promesa ASC
                """)
                alertas_retrasos = cursor.fetchall()

            # C. Repuestos Críticos (Solo para Admin/Cajero)
            if session.get('user_rol') in ['admin', 'cajero']:
                cursor.execute("""
                    SELECT id, codigo, nombre, stock_actual, stock_minimo 
                    FROM productos 
                    WHERE tipo = 'repuesto' AND stock_actual <= stock_minimo
                    ORDER BY stock_actual ASC LIMIT 6
                """)
                alertas_inventario = cursor.fetchall()

            # --- 3. GRÁFICO DE VENTAS ---
            dias_semana_es = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
            hoy_fecha = datetime.now()
            fechas_labels = []
            datos_ventas_dict = {}
            
            cursor.execute("SELECT DATE(fecha_pago) as dia, SUM(monto_total) as total FROM pagos WHERE fecha_pago >= CURDATE() - INTERVAL 6 DAY AND estado != 'anulado' GROUP BY DATE(fecha_pago)")
            ventas_db = cursor.fetchall()
            
            for venta in ventas_db:
                datos_ventas_dict[venta['dia']] = float(venta['total'])
                
            for i in range(6, -1, -1):
                fecha = hoy_fecha - timedelta(days=i)
                dia_str = dias_semana_es[fecha.weekday()]
                fechas_labels.append(f"{dia_str} {fecha.day}")
                chart_data.append(datos_ventas_dict.get(fecha.date(), 0.0))
            
            chart_labels = fechas_labels
            cursor.close()
            conn.close()
        except Error as e:
            print(f"Error Dashboard: {e}")
            
    # Formateo de fecha en español
    hoy = datetime.now()
    dias_completos = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    meses_completos = ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
    fecha_actual = f"{dias_completos[hoy.weekday()]}, {hoy.day} de {meses_completos[hoy.month - 1]} de {hoy.year}"
    
    return render_template('dashboard.html', 
                           kpis=kpis, 
                           fecha_actual=fecha_actual, 
                           chart_labels=chart_labels, 
                           chart_data=chart_data, 
                           alertas_repuestos=alertas_repuestos,
                           alertas_retrasos=alertas_retrasos,
                           alertas_inventario=alertas_inventario)

# ================== MÓDULO CATÁLOGO E INVENTARIO (ACTUALIZADO CON KPIs) ==================

@app.route('/inventario')
@roles_required(['admin'])
def inventario_index():
    conn = get_db_connection()
    productos = []
    # Añadimos 'valor_inventario' a las estadísticas
    stats = {'total_items': 0, 'bajo_stock': 0, 'total_servicios': 0, 'valor_inventario': 0.0}
    
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM productos ORDER BY tipo ASC, nombre ASC")
        productos = cursor.fetchall()
        
        cursor.execute("SELECT COUNT(*) as total FROM productos")
        stats['total_items'] = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as bajo FROM productos WHERE tipo = 'repuesto' AND stock_actual <= stock_minimo")
        stats['bajo_stock'] = cursor.fetchone()['bajo']
        
        cursor.execute("SELECT COUNT(*) as servicios FROM productos WHERE tipo = 'servicio'")
        stats['total_servicios'] = cursor.fetchone()['servicios']

        # NUEVO: Calculamos la plata total en el almacén (Stock * Costo CPP)
        cursor.execute("SELECT SUM(stock_actual * precio_compra) as plata_total FROM productos WHERE tipo = 'repuesto' AND stock_actual > 0")
        resultado_plata = cursor.fetchone()
        stats['valor_inventario'] = resultado_plata['plata_total'] if resultado_plata['plata_total'] else 0.0
        
        conn.close()
        
    return render_template('inventario/listar_productos.html', productos=productos, stats=stats)


# --- NUEVO PRODUCTO ---
@app.route('/inventario/nuevo', methods=['GET', 'POST'])
@roles_required(['admin'])
def nuevo_producto():
    conn = get_db_connection()
    if not conn: return redirect(url_for('inventario_index'))
    
    if request.method == 'POST':
        try:
            cursor = conn.cursor()
            
            # USAMOS .get() PARA EVITAR EL ERROR SI EL CAMPO ESTÁ DESHABILITADO
            marca = request.form.get('marca', '') 
            precio_compra = request.form.get('precio_compra', 0) or 0
            stock_actual = request.form.get('stock_actual', 0) or 0
            stock_minimo = request.form.get('stock_minimo', 0) or 0

            cursor.execute("""
                INSERT INTO productos (codigo, nombre, descripcion, categoria, marca, aplicacion, unidad_medida, 
                                     precio_compra, precio_venta, stock_actual, stock_minimo, tipo)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (request.form['codigo'], request.form['nombre'], request.form['descripcion'],
                  request.form['categoria'], marca, request.form['aplicacion'], request.form['unidad_medida'],
                  precio_compra, request.form['precio_venta'], 
                  stock_actual, stock_minimo, request.form['tipo']))
            conn.commit()
            flash("✅ Item registrado correctamente.", "success")
            return redirect(url_for('inventario_index'))
        except Error as e:
            flash(f"Error: {e.msg}", "danger")
        finally:
            conn.close()
    
    # Lógica para auto-generar el código correlativo
    cursor = conn.cursor()
    cursor.execute("SELECT MAX(CAST(SUBSTRING(codigo, 5) AS UNSIGNED)) FROM productos WHERE codigo LIKE 'REP-%'")
    last_rep = cursor.fetchone()[0]
    next_rep = f"REP-{(last_rep + 1):04d}" if last_rep else "REP-0001"
    
    cursor.execute("SELECT MAX(CAST(SUBSTRING(codigo, 5) AS UNSIGNED)) FROM productos WHERE codigo LIKE 'SRV-%'")
    last_srv = cursor.fetchone()[0]
    next_srv = f"SRV-{(last_srv + 1):04d}" if last_srv else "SRV-0001"
    conn.close()
    
    return render_template('inventario/nuevo_producto.html', next_rep=next_rep, next_srv=next_srv)


# --- EDITAR PRODUCTO ---
@app.route('/inventario/editar/<int:producto_id>', methods=['GET', 'POST'])
@roles_required(['admin'])
def editar_producto(producto_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('inventario_index'))
    
    if request.method == 'POST':
        try:
            cursor = conn.cursor()
            
            # --- LECTURA SEGURA DE DATOS ---
            codigo = request.form['codigo']
            nombre = request.form['nombre']
            descripcion = request.form.get('descripcion', '')
            categoria = request.form.get('categoria', '')
            unidad_medida = request.form.get('unidad_medida', 'NIU')
            precio_venta = request.form['precio_venta']
            tipo = request.form['tipo'] 
            aplicacion = request.form.get('aplicacion', 'Universal')

            # Datos opcionales (NO LEEMOS EL PRECIO_COMPRA PARA BLINDARLO)
            marca = request.form.get('marca', '') 
            stock_actual = request.form.get('stock_actual', 0) or 0
            stock_minimo = request.form.get('stock_minimo', 0) or 0

            # QUERY ACTUALIZADO SIN 'precio_compra=%s'
            cursor.execute("""
                UPDATE productos 
                SET codigo=%s, nombre=%s, descripcion=%s, categoria=%s, marca=%s, 
                    aplicacion=%s, unidad_medida=%s, precio_venta=%s, 
                    stock_actual=%s, stock_minimo=%s, tipo=%s
                WHERE id=%s
            """, (codigo, nombre, descripcion, categoria, marca, 
                  aplicacion, unidad_medida, precio_venta, 
                  stock_actual, stock_minimo, tipo, producto_id))
            
            conn.commit()
            cursor.close()
            flash("✅ Producto actualizado (El Costo Promedio se mantuvo intacto).", "info")
            return redirect(url_for('inventario_index'))
            
        except Error as e:
            # ESTA ES LA LÍNEA QUE FALTABA Y CAUSABA EL ERROR EN VS CODE
            flash(f"Error al actualizar: {e.msg}", "danger")
        finally:
            if conn.is_connected(): conn.close()
    
    # GET: Cargar datos para mostrar
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM productos WHERE id = %s", (producto_id,))
    producto = cursor.fetchone()
    conn.close()
    
    if not producto:
        flash("Producto no encontrado.", "warning")
        return redirect(url_for('inventario_index'))

    return render_template('inventario/editar_producto.html', p=producto)


# --- ELIMINAR PRODUCTO ---
@app.route('/inventario/eliminar/<int:producto_id>')
@roles_required(['admin'])
def eliminar_producto(producto_id):
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM productos WHERE id = %s", (producto_id,))
            conn.commit()
            flash("🗑️ Producto eliminado.", "success")
        except Error as e:
            if e.errno == 1451:
                flash("❌ No se puede eliminar: Este producto ya se usó en una Orden de Trabajo.", "danger")
            else:
                flash(f"Error: {e.msg}", "danger")
        finally:
            conn.close()
    return redirect(url_for('inventario_index'))

# ==============================================================================
# MÓDULO DE COMPRAS (INGRESO DE STOCK Y ACTUALIZACIÓN DE PRECIOS)
# ==============================================================================

# Configuración para guardar facturas (Asegúrate de que esto esté arriba, cerca de tus configuraciones)
UPLOAD_FOLDER_FACTURAS = 'static/uploads/facturas_compras'
app.config['UPLOAD_FOLDER_FACTURAS'] = UPLOAD_FOLDER_FACTURAS
os.makedirs(UPLOAD_FOLDER_FACTURAS, exist_ok=True)

@app.route('/inventario/compras/nueva', methods=['GET', 'POST'])
@roles_required(['admin'])
def nueva_compra():
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos.", "danger")
        return redirect(url_for('dashboard'))

    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        try:
            # 1. Capturar datos principales
            proveedor_id = request.form.get('proveedor_id')
            fecha_emision = request.form.get('fecha_emision')
            tipo_comprobante = request.form.get('tipo_comprobante', 'Factura').capitalize()
            numero_comprobante = request.form.get('numero_comprobante')
            metodo_pago = request.form.get('metodo_pago')
            
            # --- CAMPOS ENTERPRISE (Evitando errores si vienen vacíos) ---
            fecha_vencimiento = request.form.get('fecha_vencimiento')
            if not fecha_vencimiento or fecha_vencimiento.strip() == '':
                fecha_vencimiento = None 
                
            observaciones = request.form.get('observaciones', '')
            descuento = float(request.form.get('descuento', 0) or 0)
            flete = float(request.form.get('flete', 0) or 0)
            
            subtotal = float(request.form.get('subtotal', 0) or 0)
            igv = float(request.form.get('igv', 0) or 0)
            total_compra = float(request.form.get('total_compra', 0) or 0)
            
            estado_pago = 'pendiente' if metodo_pago == 'credito' else 'pagado'
            usuario_id = session.get('user_id') or session.get('id')

            # 2. Guardar Evidencia (PDF/Foto) localmente
            archivo_adjunto = None
            if 'archivo_factura' in request.files:
                file = request.files['archivo_factura']
                if file and file.filename != '':
                    filename = secure_filename(f"COMPRA_{numero_comprobante}_{file.filename}")
                    file_path = os.path.join(app.config['UPLOAD_FOLDER_FACTURAS'], filename)
                    file.save(file_path)
                    archivo_adjunto = filename

            # 3. Guardar en tabla COMPRAS
            cursor.execute("""
                INSERT INTO compras (proveedor_id, usuario_id, fecha_emision, tipo_comprobante, 
                                   numero_comprobante, subtotal, descuento, igv, flete, total_compra, 
                                   metodo_pago, fecha_vencimiento, estado_pago, observaciones, archivo_adjunto)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (proveedor_id, usuario_id, fecha_emision, tipo_comprobante, numero_comprobante, 
                  subtotal, descuento, igv, flete, total_compra, 
                  metodo_pago, fecha_vencimiento, estado_pago, observaciones, archivo_adjunto))
            
            compra_id = cursor.lastrowid 

            # 4. Detalles y Actualización de Stock/Precios
            productos_ids = request.form.getlist('producto_id[]')
            cantidades = request.form.getlist('cantidad[]')
            precios = request.form.getlist('precio[]')

            for i in range(len(productos_ids)):
                prod_id = productos_ids[i]
                cant = int(cantidades[i])
                precio_nuevo = float(precios[i])

                cursor.execute("""
                    INSERT INTO detalles_compra (compra_id, producto_id, cantidad, precio_unitario_compra)
                    VALUES (%s, %s, %s, %s)
                """, (compra_id, prod_id, cant, precio_nuevo))

                # ACTUALIZA STOCK Y EL ÚLTIMO PRECIO DE COMPRA
                # 1. Obtener el stock y precio actual del repuesto ANTES de la compra
                cursor.execute("SELECT stock_actual, precio_compra FROM productos WHERE id = %s", (prod_id,))
                prod = cursor.fetchone()
                
                stock_viejo = float(prod['stock_actual'] or 0)
                precio_viejo = float(prod['precio_compra'] or 0)
                
                # 2. Algoritmo de Costo Promedio Ponderado (CPP)
                nuevo_stock = stock_viejo + cant
                if nuevo_stock > 0:
                    nuevo_precio_promedio = ((stock_viejo * precio_viejo) + (cant * precio_nuevo)) / nuevo_stock
                else:
                    nuevo_precio_promedio = precio_nuevo
                
                # 3. Actualizar el catálogo con el Stock Real y el Costo Promedio
                cursor.execute("""
                    UPDATE productos 
                    SET stock_actual = %s,
                        precio_compra = %s
                    WHERE id = %s
                """, (nuevo_stock, nuevo_precio_promedio, prod_id))

            conn.commit()
            registrar_auditoria("Registró ingreso de almacén", "INVENTARIO", {"comprobante": numero_comprobante, "total": total_compra})
            
            flash("✅ Compra procesada: Stock sumado y costos actualizados exitosamente.", "success")
            return redirect(url_for('inventario_index')) 

        except Error as e:
            conn.rollback() 
            flash(f"Error al registrar la compra: {e.msg}", "danger")
        finally:
            cursor.close()
            conn.close()

    # --- GET: Cargar Listas ---
    cursor.execute("""
        SELECT p.id as proveedor_id, per.nombre_completo as razon_social, per.numero_documento as ruc
        FROM proveedores p
        JOIN personas per ON p.persona_id = per.id
        WHERE p.estado = 'activo'
    """)
    proveedores = cursor.fetchall()

    cursor.execute("""
        SELECT id, codigo, nombre, stock_actual, stock_minimo, precio_compra 
        FROM productos 
        WHERE tipo = 'repuesto' 
        ORDER BY nombre ASC
    """)
    productos = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('inventario/nueva_compra.html', proveedores=proveedores, productos=productos)

# ==============================================================================
# HISTORIAL DE COMPRAS (PANEL ADMINISTRATIVO)
# ==============================================================================
@app.route('/inventario/compras')
@roles_required(['admin'])
def historial_compras():
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos.", "danger")
        return redirect(url_for('dashboard'))

    cursor = conn.cursor(dictionary=True)
    
    try:
        # Extraemos todas las compras con el nombre real del proveedor
        cursor.execute("""
            SELECT c.id, c.fecha_emision, c.numero_comprobante, c.tipo_comprobante,
                   c.total_compra, c.estado_pago, c.archivo_adjunto, c.metodo_pago,
                   per.nombre_completo as razon_social
            FROM compras c
            JOIN proveedores p ON c.proveedor_id = p.id
            JOIN personas per ON p.persona_id = per.id
            ORDER BY c.fecha_emision DESC, c.id DESC
        """)
        compras = cursor.fetchall()
        
        # Cálculos rápidos para las Tarjetas (KPIs) del Dashboard
        total_invertido = sum(c['total_compra'] for c in compras)
        total_deuda = sum(c['total_compra'] for c in compras if c['estado_pago'] == 'pendiente')
        
    except Error as e:
        flash(f"Error al cargar el historial: {e.msg}", "danger")
        compras, total_invertido, total_deuda = [], 0, 0
    finally:
        cursor.close()
        conn.close()

    return render_template('inventario/historial_compras.html', 
                           compras=compras, 
                           total_invertido=total_invertido,
                           total_deuda=total_deuda)

# ==============================================================================
# API AJAX: VER DETALLES DE UNA COMPRA SIN RECARGAR LA PÁGINA
# ==============================================================================
@app.route('/api/compras/<int:compra_id>/detalles')
@roles_required(['admin'])
def api_detalles_compra(compra_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Error de conexión a la base de datos"}), 500

    cursor = conn.cursor(dictionary=True)
    try:
        # 1. Buscamos la lista de repuestos
        cursor.execute("""
            SELECT d.cantidad, d.precio_unitario_compra as precio, 
                   (d.cantidad * d.precio_unitario_compra) as subtotal,
                   p.nombre as producto, p.codigo
            FROM detalles_compra d
            JOIN productos p ON d.producto_id = p.id
            WHERE d.compra_id = %s
        """, (compra_id,))
        detalles = cursor.fetchall()
        
        # 2. Buscamos el resumen financiero de la factura (Flete, Descuentos, IGV)
        cursor.execute("""
            SELECT subtotal, descuento, igv, flete, total_compra, observaciones, fecha_vencimiento
            FROM compras
            WHERE id = %s
        """, (compra_id,))
        resumen = cursor.fetchone()

        # Enviamos ambas cosas juntas
        return jsonify({
            "items": detalles,
            "resumen": resumen
        })
    except Error as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# ==============================================================================
# LISTAR PROVEEDORES (DIRECTORIO)
# ==============================================================================
@app.route('/proveedores', methods=['GET'])
@roles_required(['admin', 'cajero'])
def listar_proveedores():
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos.", "danger")
        return redirect(url_for('dashboard'))

    cursor = conn.cursor(dictionary=True)
    # Traemos los datos uniendo la tabla proveedores con la tabla personas (Herencia)
    cursor.execute("""
        SELECT p.id, per.numero_documento, per.nombre_completo, per.telefono,
               per.email, p.nombre_contacto, p.estado
        FROM proveedores p
        JOIN personas per ON p.persona_id = per.id
        ORDER BY p.id DESC
    """)
    proveedores = cursor.fetchall()
    
    cursor.close()
    conn.close()

    return render_template('proveedores/listar_proveedores.html', proveedores=proveedores)

# ==============================================================================
# EDITAR PROVEEDOR
# ==============================================================================
@app.route('/proveedores/editar/<int:id>', methods=['GET', 'POST'])
@roles_required(['admin', 'cajero'])
def editar_proveedor(id):
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos.", "danger")
        return redirect(url_for('listar_proveedores'))

    cursor = conn.cursor(dictionary=True)

    if request.method == 'POST':
        try:
            # Capturamos datos editables
            telefono = request.form.get('telefono', '').strip()
            email = request.form.get('email', '').strip()
            direccion = request.form.get('direccion', '').strip()
            nombre_contacto = request.form.get('nombre_contacto', '').strip()
            sitio_web = request.form.get('sitio_web', '').strip()
            estado = request.form.get('estado', 'activo')

            # 1. Buscamos el ID de la persona asociada a este proveedor
            cursor.execute("SELECT persona_id FROM proveedores WHERE id = %s", (id,))
            prov = cursor.fetchone()
            
            if prov:
                persona_id = prov['persona_id']
                
                # 2. Actualizamos la tabla maestra (Personas)
                cursor.execute("""
                    UPDATE personas 
                    SET telefono = %s, email = %s, direccion = %s
                    WHERE id = %s
                """, (telefono, email, direccion, persona_id))

                # 3. Actualizamos la tabla específica (Proveedores)
                cursor.execute("""
                    UPDATE proveedores 
                    SET nombre_contacto = %s, sitio_web = %s, estado = %s
                    WHERE id = %s
                """, (nombre_contacto, sitio_web, estado, id))

                conn.commit()
                flash("✅ Proveedor actualizado con éxito.", "success")
                return redirect(url_for('listar_proveedores'))

        except Error as e:
            conn.rollback()
            flash(f"Error al actualizar la base de datos: {e.msg}", "danger")
        finally:
            cursor.close()
            conn.close()

    # GET: Cargar los datos actuales para mostrarlos en el formulario
    cursor.execute("""
        SELECT p.id, p.persona_id, p.nombre_contacto, p.sitio_web, p.estado,
               per.tipo_documento, per.numero_documento, per.nombre_completo,
               per.telefono, per.email, per.direccion
        FROM proveedores p
        JOIN personas per ON p.persona_id = per.id
        WHERE p.id = %s
    """, (id,))
    proveedor = cursor.fetchone()
    cursor.close()
    conn.close()

    if not proveedor:
        flash("Proveedor no encontrado.", "warning")
        return redirect(url_for('listar_proveedores'))

    return render_template('proveedores/editar_proveedor.html', proveedor=proveedor)

# ================== MÓDULO CLIENTES (V3.0 ARQUITECTURA SUPERTIPO & UBIGEO) ==================

@app.route('/clientes')
@roles_required(['admin', 'mecanico', 'cajero'])
def clientes_index():
    conn = get_db_connection()
    clientes = []
    stats = {'total_clientes': 0, 'total_vehiculos': 0, 'nuevos_hoy': 0}
    
    if conn:
        try:
            cursor = conn.cursor(dictionary=True)
            
            # KPI 1: Clientes Nuevos Hoy (Buscamos en la tabla personas)
            cursor.execute("""
                SELECT COUNT(*) as total FROM clientes c 
                JOIN personas p ON c.persona_id = p.id 
                WHERE DATE(p.fecha_creacion) = CURDATE()
            """)
            stats['nuevos_hoy'] = cursor.fetchone()['total']
            
            # KPI 2: Total de Vehículos
            cursor.execute("SELECT COUNT(*) as total FROM vehiculos")
            stats['total_vehiculos'] = cursor.fetchone()['total']

            # Lista principal de clientes (Cruzamos clientes con personas)
            cursor.execute("""
                SELECT c.id, p.numero_documento as dni_ruc, p.nombre_completo as nombre_razon_social, 
                       p.telefono, p.email, p.direccion, COUNT(v.id) as total_vehiculos
                FROM clientes c
                JOIN personas p ON c.persona_id = p.id
                LEFT JOIN vehiculos v ON c.id = v.cliente_id
                GROUP BY c.id, p.id
                ORDER BY p.nombre_completo ASC
            """)
            clientes = cursor.fetchall()
            stats['total_clientes'] = len(clientes) # KPI 3
            
            conn.close()
        except Error as e:
            flash(f"Error al cargar clientes: {e}", "danger")

    return render_template('clientes/listar_clientes.html', clientes=clientes, stats=stats)

@app.route('/clientes/nuevo', methods=['GET', 'POST'])
@roles_required(['admin', 'mecanico', 'cajero'])
def nuevo_cliente():
    if request.method == 'POST':
        # 1. Capturamos los datos
        tipo_doc = request.form.get('tipo_documento', 'DNI')
        numero_doc = request.form.get('numero_documento', '').strip()
        nombre = request.form.get('nombre_completo', '').strip()
        telefono = request.form.get('telefono', '').strip()
        email = request.form.get('email', '').strip()
        direccion = request.form.get('direccion', '').strip()
        departamento = request.form.get('departamento', '')
        provincia = request.form.get('provincia', '')
        distrito = request.form.get('distrito', '')
        
        # 🚨 BARRERA DE SEGURIDAD (AGREGA ESTO) 🚨
        if not numero_doc or not nombre:
            flash("❌ Bloqueo de Seguridad: El documento y el nombre son obligatorios.", "danger")
            return redirect(url_for('nuevo_cliente'))
        
        conn = get_db_connection()
        try:
            cursor = conn.cursor(dictionary=True)
            
            # ========================================================
            # LA MAGIA DE LA HERENCIA 
            # ========================================================
            # 1. ¿Esta persona (DNI/RUC) ya existe en la base maestra?
            cursor.execute("SELECT id FROM personas WHERE numero_documento = %s", (numero_doc,))
            persona_existente = cursor.fetchone()
            
            if persona_existente:
                # CASO A: ¡Es Susy o Carlos! Ya existen en la BD.
                persona_id = persona_existente['id']
                
                # Validamos que no tenga el rol de cliente duplicado
                cursor.execute("SELECT id FROM clientes WHERE persona_id = %s", (persona_id,))
                if cursor.fetchone():
                    flash("❌ Esta persona ya estaba registrada como cliente en el taller.", "warning")
                    return redirect(url_for('clientes_index'))
            else:
                # CASO B: Es un cliente de la calle. Lo insertamos desde cero en Personas.
                cursor.execute("""
                    INSERT INTO personas (tipo_documento, numero_documento, nombre_completo, telefono, email, direccion, departamento, provincia, distrito) 
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (tipo_doc, numero_doc, nombre, telefono, email, direccion, departamento, provincia, distrito))
                persona_id = cursor.lastrowid
            
            # ========================================================
            # 2. LE ASIGNAMOS LA "CAMISETA" DE CLIENTE (Usando su ID)
            # ========================================================
            cursor.execute("INSERT INTO clientes (persona_id) VALUES (%s)", (persona_id,))
            
            conn.commit()
            # --- AUDITORÍA ---
            registrar_auditoria("Registró un nuevo cliente", "CLIENTES", {"documento": numero_doc, "nombre": nombre})
            flash("✅ Cliente registrado con éxito aplicando la Arquitectura de Herencia.", "success")
            
        except Error as e:
            flash(f"Error en la base de datos: {e.msg}", "danger")
        finally:
            if conn: conn.close()
            
   # === CARGA DE UBIGEO DESDE EL SERVIDOR ===
    try:
        # Leemos el cerebro geográfico
        with open('ubigeo.json', 'r', encoding='utf-8') as f:
            ubigeo_data = json.load(f)
    except:
        ubigeo_data = {} # Si no hay archivo, mandamos vacío por seguridad
        
    return render_template('clientes/nuevo_cliente.html', ubigeo_data=json.dumps(ubigeo_data))

@app.route('/clientes/editar/<int:cliente_id>', methods=['GET', 'POST'])
@roles_required(['admin', 'cajero']) 
def editar_cliente(cliente_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            cursor = conn.cursor()
            # Actualizamos la tabla Personas a través de una subconsulta con el ID del cliente
            cursor.execute("""
                UPDATE personas 
                SET tipo_documento=%s, numero_documento=%s, nombre_completo=%s, telefono=%s, 
                    email=%s, direccion=%s, departamento=%s, provincia=%s, distrito=%s 
                WHERE id = (SELECT persona_id FROM clientes WHERE id = %s)
            """, (
                request.form.get('tipo_documento', 'DNI'), request.form['numero_documento'], 
                request.form['nombre'], request.form['telefono'], request.form.get('email', ''), 
                request.form.get('direccion', ''), request.form.get('departamento', ''),
                request.form.get('provincia', ''), request.form.get('distrito', ''), cliente_id
            ))
            conn.commit()
            flash("✅ Perfil del cliente actualizado.", "info")
            return redirect(url_for('clientes_index'))
        except Error as e:
            flash(f"Error: {e.msg}", "danger")
        finally:
            conn.close()
            
    cursor = conn.cursor(dictionary=True)
    # Seleccionamos cruzando las tablas para llenar el formulario
    cursor.execute("""
        SELECT c.id as cliente_id, p.* FROM clientes c 
        JOIN personas p ON c.persona_id = p.id 
        WHERE c.id = %s
    """, (cliente_id,))
    cliente = cursor.fetchone()
    conn.close()
    
    # Truco: Mapear los nombres de columnas para no romper tu HTML actual
    if cliente:
        cliente['dni_ruc'] = cliente['numero_documento']
        cliente['nombre_razon_social'] = cliente['nombre_completo']
        
    return render_template('clientes/editar_cliente.html', cliente=cliente)

@app.route('/clientes/eliminar/<int:cliente_id>')
@roles_required(['admin'])
def eliminar_cliente(cliente_id):
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            # Averiguamos qué persona es
            cursor.execute("SELECT persona_id FROM clientes WHERE id = %s", (cliente_id,))
            persona = cursor.fetchone()
            
            if persona:
                # Al borrar de la tabla 'clientes', mantenemos a la 'persona' por si es empleado también
                cursor.execute("DELETE FROM clientes WHERE id = %s", (cliente_id,))
                conn.commit()
                flash("🗑️ Rol de Cliente eliminado correctamente.", "success")
        except Error as e:
            if e.errno == 1451:
                flash("❌ No se puede eliminar: El cliente ya tiene vehículos u órdenes registradas.", "danger")
            else:
                flash(f"Error: {e.msg}", "danger")
        finally:
            conn.close()
    return redirect(url_for('clientes_index'))

@app.route('/clientes/perfil/<int:cliente_id>')
@roles_required(['admin', 'mecanico', 'cajero'])
def perfil_cliente(cliente_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT c.id, p.numero_documento as dni_ruc, p.nombre_completo as nombre_razon_social, 
               p.telefono, p.email, p.direccion 
        FROM clientes c 
        JOIN personas p ON c.persona_id = p.id 
        WHERE c.id = %s
    """, (cliente_id,))
    cliente = cursor.fetchone()
    
    cursor.execute("SELECT * FROM vehiculos WHERE cliente_id = %s ORDER BY placa", (cliente_id,))
    vehiculos = cursor.fetchall()
    conn.close()
    return render_template('clientes/perfil_cliente.html', cliente=cliente, vehiculos=vehiculos)

@app.route('/clientes/agregar_vehiculo/<int:cliente_id>', methods=['GET', 'POST'])
@roles_required(['admin', 'mecanico', 'cajero'])
def agregar_vehiculo(cliente_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))

    if request.method == 'POST':
        try:
            cursor = conn.cursor()
            cursor.execute("""
                INSERT INTO vehiculos (cliente_id, placa, tipo, marca, modelo, anio, color)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (cliente_id, request.form['placa'].upper(), request.form['tipo'], 
                  request.form['marca'], request.form['modelo'], 
                  request.form['anio'], request.form['color']))
            conn.commit()
            flash("✅ Vehículo agregado correctamente.", "success")
            return redirect(url_for('perfil_cliente', cliente_id=cliente_id))
        except Error as e:
            flash(f"Error: {e.msg}", "danger")
        finally:
            conn.close()

    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT c.id, p.nombre_completo as nombre_razon_social 
        FROM clientes c 
        JOIN personas p ON c.persona_id = p.id 
        WHERE c.id = %s
    """, (cliente_id,))
    cliente = cursor.fetchone()
    conn.close()
    return render_template('clientes/agregar_vehiculo.html', cliente=cliente)

# ==============================================================================
# ================== MÓDULO EXTERNO: RASTREO DE VEHÍCULOS ======================
# ==============================================================================
# Esta ruta es para que el cliente consulte desde su celular/casa.

@app.route('/', methods=['GET', 'POST'])
def estado_vehiculo():
    ordenes_encontradas = []
    mensaje_error = None

    if request.method == 'POST':
        # Capturamos un ÚNICO campo de búsqueda y lo limpiamos
        busqueda = request.form.get('busqueda', '').strip().upper()

        conn = get_db_connection()
        if conn:
            cursor = conn.cursor(dictionary=True)
            
            # 1. Si escribió PUROS NÚMEROS (DNI o RUC)
            if busqueda.isdigit():
                query = """
                    SELECT os.id as orden_id, os.estado, os.fecha_ingreso, os.fecha_promesa, 
                           os.diagnostico, v.placa, v.marca, v.modelo, p.nombre_completo
                    FROM ordenes_servicio os
                    JOIN vehiculos v ON os.vehiculo_id = v.id
                    JOIN clientes c ON v.cliente_id = c.id
                    JOIN personas p ON c.persona_id = p.id
                    WHERE p.numero_documento = %s
                    ORDER BY os.fecha_ingreso DESC
                """
                cursor.execute(query, (busqueda,))
                ordenes_encontradas = cursor.fetchall()
                error_msg = f"No se encontraron órdenes para el Documento: '{busqueda}'."

            # 2. Si tiene LETRAS Y NÚMEROS (Placa)
            else:
                query = """
                    SELECT os.id as orden_id, os.estado, os.fecha_ingreso, os.fecha_promesa, 
                           os.diagnostico, v.placa, v.marca, v.modelo, p.nombre_completo
                    FROM ordenes_servicio os
                    JOIN vehiculos v ON os.vehiculo_id = v.id
                    JOIN clientes c ON v.cliente_id = c.id
                    JOIN personas p ON c.persona_id = p.id
                    WHERE v.placa = %s
                    ORDER BY os.fecha_ingreso DESC LIMIT 1
                """
                cursor.execute(query, (busqueda,))
                orden = cursor.fetchone()
                if orden:
                    ordenes_encontradas = [orden]
                error_msg = f"No se encontró el vehículo con placa: '{busqueda}'."

            conn.close()

            if not ordenes_encontradas:
                mensaje_error = error_msg

    return render_template('clientes/consulta_cliente.html', ordenes=ordenes_encontradas, error=mensaje_error)


# ==============================================================================
# ================== MÓDULO INTERNO: GESTIÓN DE TALLER =========================
# ==============================================================================
# Estas rutas son solo para el Administrador y los Mecánicos.

@app.route('/taller')
@roles_required(['admin', 'mecanico'])
def taller_index():
    conn = get_db_connection()
    ordenes = []
    if conn:
        cursor = conn.cursor(dictionary=True)
        # DOBLE JOIN: Conectamos a Personas para sacar el nombre del cliente y del mecánico
        sql = """
            SELECT os.*, v.placa, v.tipo, v.marca, v.modelo, 
                   pc.nombre_completo as cliente, 
                   pu.nombre_completo as mecanico
            FROM ordenes_servicio os
            JOIN vehiculos v ON os.vehiculo_id = v.id
            JOIN clientes c ON v.cliente_id = c.id
            JOIN personas pc ON c.persona_id = pc.id
            LEFT JOIN usuarios u ON os.usuario_id = u.id
            LEFT JOIN personas pu ON u.persona_id = pu.id
            ORDER BY 
                CASE 
                    WHEN os.estado = 'esperando_repuesto' THEN 1
                    WHEN os.estado = 'pendiente' THEN 2 
                    WHEN os.estado = 'en_proceso' THEN 3 
                    WHEN os.estado = 'finalizado' THEN 4 
                    WHEN os.estado = 'entregado' THEN 5 
                END ASC, 
                os.fecha_ingreso ASC
        """
        cursor.execute(sql)
        ordenes = cursor.fetchall()
        conn.close()
    return render_template('taller/listar_ordenes.html', ordenes=ordenes)

@app.route('/taller/nueva', methods=['GET', 'POST'])
@roles_required(['admin', 'mecanico'])
def nueva_orden():
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            fp = request.form['fecha_promesa'] if request.form['fecha_promesa'] else None
            cursor = conn.cursor()
            cursor.execute("INSERT INTO ordenes_servicio (vehiculo_id, usuario_id, falla_reportada, estado, fecha_promesa) VALUES (%s, %s, %s, 'pendiente', %s)", 
                           (request.form['vehiculo_id'], request.form['usuario_id'], request.form['falla_reportada'], fp))
            conn.commit()
            flash("✅ Orden creada exitosamente.", "success")
            return redirect(url_for('taller_index'))
        except Error as e:
            flash(f"Error: {e.msg}", "danger")
        finally:
            conn.close()
            
    cursor = conn.cursor(dictionary=True)
    # JOIN para traer los vehículos con el nombre real de su dueño
    cursor.execute("""
        SELECT v.id, v.placa, v.marca, v.modelo, p.nombre_completo as cliente 
        FROM vehiculos v 
        JOIN clientes c ON v.cliente_id = c.id 
        JOIN personas p ON c.persona_id = p.id
        ORDER BY v.placa
    """)
    vehiculos = cursor.fetchall()
    
    # JOIN para traer a los mecánicos activos
    cursor.execute("""
        SELECT u.id, p.nombre_completo as nombre 
        FROM usuarios u 
        JOIN personas p ON u.persona_id = p.id 
        WHERE u.rol = 'mecanico' AND u.activo = 1
    """)
    mecanicos = cursor.fetchall()
    conn.close()
    return render_template('taller/nueva_orden.html', vehiculos=vehiculos, mecanicos=mecanicos)

@app.route('/taller/orden/<int:orden_id>', methods=['GET', 'POST'])
@roles_required(['admin', 'mecanico'])
def detalle_orden(orden_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        accion = request.form.get('accion')
        
        # INICIAR TRABAJO
        if accion == 'iniciar_trabajo':
            try:
                cursor = conn.cursor()
                cursor.execute("UPDATE ordenes_servicio SET estado = 'en_proceso' WHERE id = %s", (orden_id,))
                conn.commit()
                cursor.close()
                flash("🔧 Revisión iniciada. El vehículo ha pasado a 'En Taller'.", "info")
                return redirect(url_for('detalle_orden', orden_id=orden_id))
            except Error as e:
                flash(f"Error al iniciar trabajo: {e.msg}", "danger")

        # =====================================================================
        # AGREGAR REPUESTO/SERVICIO (VERSIÓN ENTERPRISE CON COSTEO)
        # =====================================================================
        elif accion == 'agregar_item':
            try:
                pid = request.form['producto_id']
                cant = int(request.form['cantidad'])
                cursor = conn.cursor(dictionary=True)
                
                # EXTRAEMOS TAMBIÉN EL PRECIO_COMPRA (Costo Promedio)
                cursor.execute("SELECT precio_venta, precio_compra, stock_actual, tipo FROM productos WHERE id = %s", (pid,))
                prod = cursor.fetchone()
                
                if prod:
                    # BARRERA ANTI-FANTASMAS MEJORADA
                    if prod['tipo'] == 'repuesto' and prod['stock_actual'] < cant:
                        flash(f"❌ Stock insuficiente. Solo quedan {prod['stock_actual']} en almacén.", "danger")
                    else:
                        # 1. Definimos el costo real (Si es servicio, el costo base es 0)
                        costo_real = prod['precio_compra'] if prod['tipo'] == 'repuesto' else 0.00
                        
                        # 2. Guardamos el Precio de Venta Y el Costo Promedio
                        cursor.execute("""
                            INSERT INTO detalles_orden (orden_id, producto_id, cantidad, precio_unitario, costo_unitario) 
                            VALUES (%s, %s, %s, %s, %s)
                        """, (orden_id, pid, cant, prod['precio_venta'], costo_real))
                        
                        # 3. Descontamos del Almacén en tiempo real
                        if prod['tipo'] == 'repuesto':
                            cursor.execute("UPDATE productos SET stock_actual = stock_actual - %s WHERE id = %s", (cant, pid))
                        
                        conn.commit()
                        
                        # 4. Auditoría Silenciosa
                        registrar_auditoria("Despachó item a Orden de Trabajo", "TALLER", {"orden_id": orden_id, "producto": pid, "cant": cant})
                        
                        flash("✅ Repuesto/Servicio despachado y costeado correctamente.", "success")
                cursor.close()
            except Error as e:
                flash(f"Error al agregar item: {e.msg}", "danger")

        # FINALIZAR ORDEN
        elif accion == 'finalizar_orden':
            try:
                informe_final = request.form.get('diagnostico_final')
                cursor = conn.cursor(dictionary=True) # Usamos dictionary para extraer datos más fácil

                # 1. Buscamos el correo, nombre y placa ANTES de actualizar
                cursor.execute("""
                    SELECT v.placa, pc.email, pc.nombre_completo as cliente
                    FROM ordenes_servicio os
                    JOIN vehiculos v ON os.vehiculo_id = v.id
                    JOIN clientes c ON v.cliente_id = c.id
                    JOIN personas pc ON c.persona_id = pc.id
                    WHERE os.id = %s
                """, (orden_id,))
                datos_cliente = cursor.fetchone()

                # 2. Actualizamos la orden en la base de datos
                cursor.execute("""
                    UPDATE ordenes_servicio 
                    SET estado = 'finalizado', diagnostico = %s 
                    WHERE id = %s
                """, (informe_final, orden_id))
                conn.commit()
                
                # --- AUDITORÍA ---
                registrar_auditoria("Finalizó reparación de vehículo", "TALLER", {"orden_id": orden_id, "diagnostico": informe_final[:50]})

                # 3. Enviamos el correo automáticamente si tiene uno registrado
                if datos_cliente and datos_cliente.get('email'):
                    enviado = enviar_correo_vehiculo_listo(
                        correo_cliente=datos_cliente['email'],
                        nombre_cliente=datos_cliente['cliente'],
                        placa=datos_cliente['placa'],
                        orden_id=orden_id,
                        informe_ia=informe_final
                    )
                    if enviado:
                        flash("🎉 Orden finalizada y cliente notificado por correo electrónico.", "success")
                    else:
                        flash("🎉 Orden finalizada (No se pudo conectar con Gmail).", "warning")
                else:
                    flash("🎉 Orden finalizada (El cliente no tiene correo registrado).", "success")

                cursor.close()
                return redirect(url_for('taller_index'))
            except Error as e:
                flash(f"Error al finalizar: {e.msg}", "danger")
    
    # --- GET: CARGAR DATOS PARA MOSTRAR LA ORDEN ---
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT os.*, v.placa, v.marca, v.modelo, 
               pc.nombre_completo as cliente, 
               pu.nombre_completo as mecanico 
        FROM ordenes_servicio os 
        JOIN vehiculos v ON os.vehiculo_id = v.id 
        JOIN clientes c ON v.cliente_id = c.id 
        JOIN personas pc ON c.persona_id = pc.id
        LEFT JOIN usuarios u ON os.usuario_id = u.id 
        LEFT JOIN personas pu ON u.persona_id = pu.id
        WHERE os.id = %s
    """, (orden_id,))
    orden = cursor.fetchone()
    
    cursor.execute("""
        SELECT do.*, p.nombre as producto, p.codigo, p.tipo
        FROM detalles_orden do 
        JOIN productos p ON do.producto_id = p.id 
        WHERE do.orden_id = %s
    """, (orden_id,))
    detalles = cursor.fetchall()
    
    # Productos Disponibles
    cursor.execute("SELECT * FROM productos WHERE stock_actual > 0 OR tipo = 'servicio' ORDER BY nombre")
    prods = cursor.fetchall()
    
    # =======================================================
    # NUEVO: PRODUCTOS AGOTADOS (Para el modal inteligente)
    # =======================================================
    cursor.execute("SELECT * FROM productos WHERE tipo = 'repuesto' AND stock_actual <= 0 ORDER BY nombre")
    prods_agotados = cursor.fetchall()
    
    cursor.close()
    conn.close()

    total_acumulado = sum(d['subtotal'] for d in detalles if d['subtotal'] is not None)
    total_float = float(total_acumulado)
    subtotal = total_float / 1.18
    igv = total_float - subtotal

    return render_template('taller/detalle_orden.html', 
                           orden=orden, detalles=detalles, productos=prods,
                           productos_agotados=prods_agotados, # <-- Aquí pasamos la nueva variable al HTML
                           total=total_float, subtotal=subtotal, igv=igv)

@app.route('/taller/editar/<int:orden_id>', methods=['GET', 'POST'])
@roles_required(['admin', 'mecanico'])
def editar_orden(orden_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            cursor = conn.cursor()
            cursor.execute("UPDATE ordenes_servicio SET usuario_id = %s, falla_reportada = %s, fecha_promesa = %s WHERE id = %s AND estado NOT IN ('finalizado', 'entregado')",
                           (request.form['usuario_id'], request.form['falla_reportada'], request.form['fecha_promesa'] if request.form['fecha_promesa'] else None, orden_id))
            conn.commit()
            flash("✅ Orden actualizada.", "info")
            return redirect(url_for('taller_index'))
        except Error as e:
            flash(f"Error: {e.msg}", "danger")
    
    cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM ordenes_servicio WHERE id = %s", (orden_id,))
    orden = cursor.fetchone()
    
    if orden['estado'] in ['finalizado', 'entregado']:
        flash("⚠️ No se pueden editar órdenes finalizadas.", "warning")
        return redirect(url_for('taller_index'))
        
    cursor.execute("SELECT v.placa, v.marca, v.modelo FROM vehiculos v WHERE id = %s", (orden['vehiculo_id'],))
    vehiculo = cursor.fetchone()
    
    # JOIN para mecánicos
    cursor.execute("""
        SELECT u.id, p.nombre_completo as nombre 
        FROM usuarios u 
        JOIN personas p ON u.persona_id = p.id 
        WHERE u.rol = 'mecanico' AND u.activo = 1
    """)
    mecanicos = cursor.fetchall()
    conn.close()
    return render_template('taller/editar_orden.html', orden=orden, vehiculo=vehiculo, mecanicos=mecanicos)

@app.route('/taller/orden/eliminar_item/<int:detalle_id>')
@roles_required(['admin', 'mecanico'])
def eliminar_detalle_orden(detalle_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    try:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT do.orden_id, do.producto_id, do.cantidad, p.tipo 
            FROM detalles_orden do
            JOIN productos p ON do.producto_id = p.id
            WHERE do.id = %s
        """, (detalle_id,))
        item = cursor.fetchone()
        
        if item:
            if item['tipo'] == 'repuesto':
                cursor.execute("UPDATE productos SET stock_actual = stock_actual + %s WHERE id = %s", 
                               (item['cantidad'], item['producto_id']))
            
            cursor.execute("DELETE FROM detalles_orden WHERE id = %s", (detalle_id,))
            conn.commit()
            flash("🗑️ Item eliminado (Stock restaurado).", "success")
            return redirect(url_for('detalle_orden', orden_id=item['orden_id']))
            
    except Error as e:
        flash(f"Error al eliminar: {e.msg}", "danger")
    finally:
        conn.close()
        
    return redirect(url_for('taller_index'))

@app.route('/taller/anular/<int:orden_id>')
@roles_required(['admin', 'mecanico'])
def anular_orden(orden_id):
    conn = get_db_connection()
    if conn:
        try:
            cursor = conn.cursor()
            cursor.execute("DELETE FROM ordenes_servicio WHERE id = %s AND estado NOT IN ('finalizado', 'entregado')", (orden_id,))
            if cursor.rowcount > 0:
                conn.commit()
                flash("🗑️ Orden anulada.", "success")
            else:
                flash("⚠️ No se puede anular una orden ya finalizada.", "warning")
        except Error as e:
            if e.errno == 1451:
                flash("❌ No se puede anular: La orden ya tiene repuestos o pagos.", "danger")
            else:
                flash(f"Error: {e.msg}", "danger")
        finally:
            conn.close()
    return redirect(url_for('taller_index'))


# ==============================================================================
# ================== ASISTENTE DE IA (REDACTOR DE CORREOS) =====================
# ==============================================================================

@app.route('/api/ia/diagnostico', methods=['POST'])
@roles_required(['admin', 'mecanico'])
def ia_diagnostico():
    data = request.get_json()
    diagnostico_mecanico = data.get('falla', '') 
    
    # NUEVO: Atrapamos el nombre y la placa que el HTML nos está enviando
    nombre_cliente = data.get('cliente', 'Estimado Cliente')
    placa_vehiculo = data.get('vehiculo', 'su vehículo')
    
    if not diagnostico_mecanico:
        return jsonify({"error": "El mecánico debe escribir primero su diagnóstico para que la IA pueda redactar el correo."}), 400
    
    # NUEVA INSTRUCCIÓN (PROMPT) BLINDADA PARA LA IA
    prompt = f"""
    Eres el asistente virtual de atención al cliente del taller automotriz 'Multiservicios Virgen de Chapi'.
    El mecánico ha finalizado de revisar el vehículo con placa {placa_vehiculo} y ha escrito las siguientes notas técnicas: "{diagnostico_mecanico}".
    
    Tu tarea es redactar el cuerpo de un correo electrónico formal, cordial y empático dirigido EXCLUSIVAMENTE a: {nombre_cliente}.
    
    REGLAS ESTRICTAS:
    1. Inicia con un saludo formal usando exactamente el nombre: {nombre_cliente}. 
    2. ¡ESTÁ TERMINANTEMENTE PROHIBIDO usar corchetes [ ] o dejar espacios en blanco para rellenar!
    3. Explica claramente el diagnóstico y los trabajos realizados, basándote ÚNICAMENTE en las notas del mecánico (NO inventes fallas ni repuestos).
    4. Indica que el vehículo {placa_vehiculo} ya se encuentra finalizado y listo para ser recogido.
    5. Despídete cordialmente a nombre del equipo de Multiservicios Virgen de Chapi.
    
    No incluyas el campo de 'Asunto', solo redacta el texto del mensaje listo para ser enviado.
    """
    
    try:
        modelo_disponible = 'gemini-1.5-flash' 
        for m in genai.list_models():
            if 'generateContent' in m.supported_generation_methods and 'gemini' in m.name.lower():
                modelo_disponible = m.name
                break 
                
        model = genai.GenerativeModel(modelo_disponible)
        response = model.generate_content(prompt)
        
        return jsonify({"diagnostico": response.text})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ==============================================================================
# IA: ESCANEO BIMODAL (SOPORTA PDF NATIVO Y FOTOS)
# ==============================================================================
@app.route('/api/ia/escanear_factura', methods=['POST'])
@roles_required(['admin'])
def ia_escanear_factura():
    if 'archivo_factura' not in request.files:
        return jsonify({"error": "No hay archivo"}), 400
    
    file = request.files['archivo_factura']
    try:
        # Guardado seguro
        temp_path = os.path.join(tempfile.gettempdir(), secure_filename(file.filename))
        file.save(temp_path)
        
        # Usamos el modelo más estable y común
        model = genai.GenerativeModel('gemini-pro') 
        
        # Leemos el archivo como binario (universal)
        with open(temp_path, "rb") as f:
            archivo_data = f.read()
            
        prompt = """Extrae RUC, razon social, numero comprobante y productos en JSON: 
        {"ruc_proveedor":"", "razon_social":"", "tipo_comprobante":"", "numero_comprobante":"", "productos": [{"descripcion":"", "cantidad":1, "precio_unitario":0.0}]}"""
        
        response = model.generate_content([prompt, {"mime_type": file.mimetype, "data": archivo_data}])
        
        texto = response.text.replace('```json', '').replace('```', '').strip()
        os.remove(temp_path)
        return jsonify(json.loads(texto))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# ================== FUNCIÓN: AVISO DE VEHÍCULO LISTO ==================
def enviar_correo_vehiculo_listo(correo_cliente, nombre_cliente, placa, orden_id, informe_ia):
    if not correo_cliente or correo_cliente.strip() == "":
        return False

    MI_CORREO = "arturo20010803@gmail.com" 
    MI_PASSWORD = "wxvqbwoflnoojhqh"
    
    mensaje = MIMEMultipart()
    mensaje['From'] = f"Taller Virgen de Chapi <{MI_CORREO}>"
    mensaje['To'] = correo_cliente
    mensaje['Subject'] = f"✅ ¡Tu vehículo {placa} está listo! - Orden OT-{orden_id:04d}"

    cuerpo_html = f"""
    <html>
    <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f6f9; padding: 20px;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
            <div style="background-color: #2563eb; color: white; padding: 20px; text-align: center;">
                <h2 style="margin: 0; font-size: 24px;">Taller Virgen de Chapi</h2>
            </div>
            
            <div style="padding: 30px;">
                <h3 style="color: #1e293b; margin-top: 0;">¡Hola, {nombre_cliente}!</h3>
                <p style="color: #475569; font-size: 16px; line-height: 1.6;">
                    Te informamos que los servicios técnicos de tu vehículo con placa <strong style="color: #2563eb;">{placa}</strong> 
                    han sido concluidos exitosamente. Ya puedes pasar a recogerlo por nuestras instalaciones.
                </p>
                
                <div style="background-color: #f8fafc; border-left: 4px solid #10b981; padding: 15px; margin: 25px 0; border-radius: 0 8px 8px 0;">
                    <h4 style="margin: 0 0 10px 0; color: #10b981;">Resumen del Trabajo Realizado:</h4>
                    <p style="margin: 0; color: #334155; font-size: 14px; white-space: pre-wrap;">{informe_ia}</p>
                </div>
                
                <p style="color: #475569; font-size: 15px;">Por favor, acércate a nuestra área de caja indicando tu número de orden: <strong>OT-{orden_id:04d}</strong>.</p>
                
                <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #e2e8f0; text-align: center;">
                    <p style="color: #94a3b8; font-size: 13px; margin: 0;">Gracias por confiar en nosotros.</p>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    mensaje.attach(MIMEText(cuerpo_html, 'html'))

    try:
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login(MI_CORREO, MI_PASSWORD)
        servidor.send_message(mensaje)
        servidor.quit()
        return True
    except Exception as e:
        print(f"Error enviando correo de terminación: {e}")
        return False

# ==============================================================================
# MÓDULO CAJA (Apertura, Pendientes, Cobros y Cierre)
# ==============================================================================

@app.route('/caja/apertura', methods=['GET', 'POST'])
@roles_required(['admin', 'cajero'])
def apertura_caja():
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
        
    cur = conn.cursor(dictionary=True)
    usuario_id = session.get('user_id') or session.get('id')
    
    # 1. Verificar si ya tiene caja abierta
    cur.execute("SELECT id FROM caja_sesiones WHERE usuario_id = %s AND estado = 'abierta'", (usuario_id,))
    if cur.fetchone():
        flash('Ya tienes una sesión de caja abierta.', 'warning')
        cur.close()
        conn.close()
        return redirect(url_for('caja_index'))
        
    # 2. Registrar Apertura
    if request.method == 'POST':
        try:
            monto_inicial = float(request.form['monto_inicial'])
            observaciones = request.form.get('observaciones', '')
            
            cur.execute("""
                INSERT INTO caja_sesiones (usuario_id, monto_inicial, estado, observaciones) 
                VALUES (%s, %s, 'abierta', %s)
            """, (usuario_id, monto_inicial, observaciones))
            conn.commit()
            
            flash('¡Caja abierta exitosamente! Ya puedes registrar cobros.', 'success')
            return redirect(url_for('caja_index'))
            
        except Error as e:
            flash(f'Error al abrir la caja: {e}', 'danger')
        finally:
            cur.close()
            conn.close()
            
    if conn.is_connected():
        cur.close()
        conn.close()
        
    fecha_hoy = datetime.now().strftime("%d/%m/%Y")
    return render_template('caja/apertura_caja.html', fecha_hoy=fecha_hoy)

# ==============================================================================
# Motor de Conexión SUNAT
# ==============================================================================

def enviar_comprobante_sunat(pago_id, nro_comprobante, tipo_comprobante, cliente, detalles, totales):
    """
    Se comunica con el proveedor de facturación (PSE) enviando un JSON.
    Devuelve la respuesta de SUNAT (Hash, Enlaces PDF/XML, Estado).
    """
    tipo_doc_sunat = 1 if tipo_comprobante == 'factura' else 2
    tipo_doc_cliente = 6 if len(cliente['dni_ruc']) == 11 else 1
    
    # Separar la serie y el número (Ej: de "B001-00000016" a "B001" y "16")
    serie, numero = nro_comprobante.split('-')
    
    # Construcción del Payload estándar para facturación electrónica
    payload = {
        "operacion": "generar_comprobante",
        "tipo_de_comprobante": tipo_doc_sunat,
        "serie": serie,
        "numero": int(numero),
        "sunat_transaction": 1,
        "cliente_tipo_de_documento": tipo_doc_cliente,
        "cliente_numero_de_documento": cliente['dni_ruc'],
        "cliente_denominacion": cliente['nombre_completo'],
        "cliente_direccion": cliente.get('direccion', 'Dirección no registrada'),
        "cliente_email": cliente.get('email', ''),
        "fecha_de_emision": datetime.now().strftime('%Y-%m-%d'),
        "moneda": 1, # 1 = Soles
        "porcentaje_de_igv": 18.00,
        "descuento_global": "",
        "total_gravada": totales['subtotal'],
        "total_igv": totales['igv'],
        "total": totales['total'],
        "detraccion": False,
        "enviar_automaticamente_a_la_sunat": True,
        "enviar_automaticamente_al_cliente": True,
        "items": []
    }
    
    # Llenar los ítems (Repuestos y Servicios)
    for item in detalles:
        item_data = {
            "unidad_de_medida": "NIU" if item.get('tipo') == 'repuesto' else "ZZ",
            "codigo": item.get('codigo', 'S/C'),
            "descripcion": item['nombre'],
            "cantidad": item['cantidad'],
            "valor_unitario": float(item['precio_unitario']) / 1.18, # Precio sin IGV
            "precio_unitario": float(item['precio_unitario']),       # Precio con IGV
            "subtotal": (float(item['cantidad']) * float(item['precio_unitario'])) / 1.18,
            "tipo_de_igv": 1, # 1 = Gravado - Operación Onerosa
            "igv": (float(item['cantidad']) * float(item['precio_unitario'])) - ((float(item['cantidad']) * float(item['precio_unitario'])) / 1.18),
            "total": float(item['cantidad']) * float(item['precio_unitario']),
            "anticipo_regularizacion": False
        }
        payload["items"].append(item_data)
        
    headers = {
        "Authorization": f"Bearer {API_FACTURACION_TOKEN}",
        "Content-Type": "application/json"
    }
    
    try:
        respuesta = requests.post(API_FACTURACION_URL, json=payload, headers=headers, timeout=10)
        datos_respuesta = respuesta.json()
        
        if respuesta.status_code == 200 and 'enlace' in datos_respuesta:
            return {
                "exito": True,
                "estado_sunat": "aceptado",
                "sunat_hash": datos_respuesta.get('codigo_hash', ''),
                "enlace_pdf": datos_respuesta.get('enlace_del_pdf', ''),
                "enlace_xml": datos_respuesta.get('enlace_del_xml', '')
            }
        else:
            return {"exito": False, "error": datos_respuesta.get('errors', 'Error desconocido en el PSE')}
            
    except requests.exceptions.RequestException as e:
        return {"exito": False, "error": f"Error de conexión con el facturador: {str(e)}"}

# ==============================================================================
# INICIO DE FUCIONES DE CAJA
# ==============================================================================
@app.route('/caja')
@roles_required(['admin', 'cajero'])
def caja_index():
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    por_cobrar = []
    cobrados_hoy = []
    usuario_id = session.get('user_id') or session.get('id')
    
    try:
        # buffered=True evita el error de la base de datos "Unread result found"
        cursor = conn.cursor(dictionary=True, buffered=True)
        
        # 1. EL GUARDIA: ¿Tiene la caja abierta actualmente?
        cursor.execute("SELECT id FROM caja_sesiones WHERE usuario_id = %s AND estado = 'abierta'", (usuario_id,))
        sesion_caja = cursor.fetchone()
        
        if not sesion_caja:
            # 2. EL AUDITOR: ¿Ya cerró una caja el día de HOY?
            cursor.execute("SELECT id, descuadre FROM caja_sesiones WHERE usuario_id = %s AND estado = 'cerrada' AND DATE(fecha_cierre) = CURDATE()", (usuario_id,))
            caja_cerrada_hoy = cursor.fetchone()

            if caja_cerrada_hoy:
                # ====================================================================
                # NUEVA LÓGICA: Traer todo el resumen para el Reporte Final (CORREGIDA)
                # ====================================================================
                cursor.execute("""
                    SELECT 
                        cs.*, 
                        (SELECT COALESCE(SUM(monto), 0) FROM caja_movimientos WHERE caja_sesion_id = cs.id AND metodo_pago != 'efectivo') as total_digital
                    FROM caja_sesiones cs
                    WHERE cs.id = %s
                """, (caja_cerrada_hoy['id'],))
                reporte_completo = cursor.fetchone()
                
                # Inyectamos el nombre directamente desde la sesión para evitar errores SQL
                reporte_completo['nombre_usuario'] = session.get('user_nombre') or session.get('nombre') or 'Cajero'

                # Mostrar pantalla bloqueada (El nuevo diseño de Reporte)
                cursor.close()
                conn.close()
                return render_template('caja/turno_finalizado.html', caja=reporte_completo)

            # 3. Si no ha abierto ni cerrado hoy, mandarlo a abrir
            flash("Debes realizar la apertura de caja (fondo fijo) antes de empezar a cobrar.", "warning")
            cursor.close()
            conn.close()
            return redirect(url_for('apertura_caja'))
        
        # 4. CARGAR PANTALLA PRINCIPAL (Órdenes pendientes)
        cursor.execute("""
            SELECT os.id, os.fecha_ingreso, 
                   p.nombre_completo as cliente, 
                   v.placa, v.modelo, 
                   SUM(do.cantidad * do.precio_unitario) as total_pagar
            FROM ordenes_servicio os 
            JOIN vehiculos v ON os.vehiculo_id = v.id 
            JOIN clientes c ON v.cliente_id = c.id
            JOIN personas p ON c.persona_id = p.id
            LEFT JOIN detalles_orden do ON os.id = do.orden_id 
            LEFT JOIN pagos pg ON os.id = pg.orden_id
            WHERE os.estado = 'finalizado' AND pg.id IS NULL 
            GROUP BY os.id 
            ORDER BY os.fecha_ingreso DESC
        """)
        por_cobrar = cursor.fetchall()
        
        # 5. CARGAR HISTORIAL CON FILTRO DE FECHAS (Inicio y Fin)
        fecha_inicio = request.args.get('fecha_inicio')
        fecha_fin = request.args.get('fecha_fin')
        
        query_historial = """
            SELECT pg.id as pago_id, pg.monto_total, pg.comprobante, pg.tipo_comprobante, pg.estado,
                   os.id as orden_id, p.nombre_completo as cliente, pg.fecha_pago
            FROM pagos pg 
            JOIN ordenes_servicio os ON pg.orden_id = os.id
            JOIN vehiculos v ON os.vehiculo_id = v.id 
            JOIN clientes c ON v.cliente_id = c.id
            JOIN personas p ON c.persona_id = p.id
        """
        params_historial = []

        if fecha_inicio and fecha_fin:
            query_historial += " WHERE DATE(pg.fecha_pago) BETWEEN %s AND %s"
            params_historial.extend([fecha_inicio, fecha_fin])
        else:
            # Si no hay fechas seleccionadas, mostramos solo los de hoy por defecto
            query_historial += " WHERE DATE(pg.fecha_pago) = CURDATE()"

        query_historial += " ORDER BY pg.fecha_pago DESC"

        cursor.execute(query_historial, tuple(params_historial))
        cobrados_hoy = cursor.fetchall()

# 6. SUPERVISIÓN DE CAJAS (SOLO PARA ADMIN)
        cajas_supervision = []
        if session.get('user_rol') == 'admin':
            cursor.execute("""
                SELECT cs.id, cs.estado, cs.monto_inicial, cs.monto_calculado, cs.fecha_apertura, cs.fecha_cierre,
                       p.nombre_completo as cajero
                FROM caja_sesiones cs
                JOIN usuarios u ON cs.usuario_id = u.id
                JOIN personas p ON u.persona_id = p.id
                ORDER BY CASE WHEN cs.estado = 'abierta' THEN 1 ELSE 2 END, cs.fecha_apertura DESC
                LIMIT 50
            """)
            cajas_supervision = cursor.fetchall()

    except Error as e:
        flash(f"Error en caja: {e}", "danger")
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
            
    return render_template('caja/pendientes.html', por_cobrar=por_cobrar, cobrados_hoy=cobrados_hoy, cajas_supervision=cajas_supervision)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Response
import io
from datetime import datetime

@app.route('/exportar_ventas_excel')
@roles_required(['admin', 'cajero'])
def exportar_ventas_excel():
    # Atrapamos las fechas si el usuario usó el filtro
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')
    
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos.", "danger")
        return redirect(url_for('vista_caja')) # O la ruta de tu caja
        
    cursor = conn.cursor(dictionary=True)
    
    # Consulta a la base de datos
    query = """
        SELECT pg.fecha_pago, pg.comprobante, pg.tipo_comprobante, 
               p.nombre_completo as cliente, pg.monto_total, pg.estado
        FROM pagos pg 
        JOIN ordenes_servicio os ON pg.orden_id = os.id
        JOIN vehiculos v ON os.vehiculo_id = v.id 
        JOIN clientes c ON v.cliente_id = c.id
        JOIN personas p ON c.persona_id = p.id
    """
    params = []
    
    if fecha_inicio and fecha_fin:
        query += " WHERE DATE(pg.fecha_pago) BETWEEN %s AND %s"
        params.extend([fecha_inicio, fecha_fin])
    else:
        query += " WHERE DATE(pg.fecha_pago) = CURDATE()"
        
    query += " ORDER BY pg.fecha_pago DESC"
    
    cursor.execute(query, tuple(params))
    ventas = cursor.fetchall()
    cursor.close()
    conn.close()

    # 1. Crear el libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Auditoría de Ventas"

    # 2. Definir Estilos
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid") # Color oscuro moderno
    header_font = Font(color="FFFFFF", bold=True)
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(horizontal="center", vertical="center")
    
    font_anulado = Font(color="9C0006", bold=True)
    fill_anulado = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # 3. Escribir Título Principal
    ws.merge_cells('A1:E1')
    ws['A1'] = "REPORTE DE AUDITORÍA DE VENTAS - MULTISERVICIOS VIRGEN DE CHAPI"
    ws['A1'].font = Font(size=14, bold=True, color="0F172A")
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:E2')
    periodo = f"Periodo: {fecha_inicio} al {fecha_fin}" if fecha_inicio else f"Periodo: {datetime.now().strftime('%Y-%m-%d')} (Hoy)"
    ws['A2'] = periodo
    ws['A2'].alignment = Alignment(horizontal="center")

    # 4. Escribir Encabezados de la tabla (Fila 4)
    headers = ['FECHA Y HORA', 'COMPROBANTE', 'CLIENTE', 'ESTADO', 'MONTO TOTAL']
    ws.append([]) # Fila 3 vacía
    ws.append(headers) # Fila 4 con encabezados

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        cell.border = border_thin

    # 5. Insertar los Datos
    total_recaudado = 0
    for row_num, venta in enumerate(ventas, 5):
        # Formatear comprobante (ej: B001-00000016 BOLETA)
        comp_texto = f"{venta['comprobante']} {str(venta['tipo_comprobante']).upper()}"
        
        row_data = [
            venta['fecha_pago'].strftime('%d/%m/%Y %I:%M %p'),
            comp_texto,
            venta['cliente'],
            str(venta['estado']).upper(),
            venta['monto_total']
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.border = border_thin
            
            # Centrar Fecha y Comprobante
            if col_num in [1, 2, 4]:
                cell.alignment = align_center
                
            # Formato de Moneda para el Monto
            if col_num == 5:
                cell.number_format = '"S/" #,##0.00'
                if venta['estado'] != 'anulado':
                    total_recaudado += float(venta['monto_total'])
            
            # Si está anulado, pintar de rojo
            if venta['estado'] == 'anulado':
                cell.font = font_anulado
                cell.fill = fill_anulado

    # 6. Fila de Total
    ultima_fila = len(ventas) + 5
    ws.merge_cells(f'A{ultima_fila}:D{ultima_fila}')
    cell_total_text = ws.cell(row=ultima_fila, column=1)
    cell_total_text.value = "TOTAL RECAUDADO (VÁLIDOS):"
    cell_total_text.font = Font(bold=True)
    cell_total_text.alignment = Alignment(horizontal="right")
    
    cell_total_monto = ws.cell(row=ultima_fila, column=5)
    cell_total_monto.value = total_recaudado
    cell_total_monto.number_format = '"S/" #,##0.00'
    cell_total_monto.font = Font(bold=True)
    cell_total_monto.border = border_thin

    # 7. Ajustar anchos de columna automáticamente
    column_widths = {'A': 22, 'B': 30, 'C': 40, 'D': 15, 'E': 20}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 8. Guardar en memoria y enviar al usuario
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Auditoria_Ventas_{datetime.now().strftime('%d-%m-%Y')}.xlsx"
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename={filename}'}
    )

import time
import hashlib
import uuid

# ==============================================================================
# MOTOR FACTURADOR - MODO SIMULACIÓN (PRUEBAS LOCALES)
# ==============================================================================
def enviar_comprobante_sunat(pago_id, nro_comprobante, tipo_comprobante, cliente, detalles, totales):
    """
    SIMULADOR: Finge la conexión con un PSE/SUNAT para pruebas de desarrollo.
    Genera un Hash aleatorio y simula una respuesta exitosa.
    """
    # 1. Simulamos el tiempo de espera de internet (1.5 segundos)
    time.sleep(1.5)
    
    # 2. Generamos un "Hash SUNAT" falso pero con el formato real que usa la SUNAT (Termina en '=')
    hash_falso = hashlib.sha256(str(uuid.uuid4()).encode()).hexdigest()[:28] + "="
    
    # 3. Devolvemos una respuesta de éxito tal cual lo haría Nubefact o Apis.net.pe
    return {
        "exito": True,
        "estado_sunat": "aceptado",
        "sunat_hash": hash_falso,
        "enlace_pdf": f"https://api.tu-proveedor.com/pdf/demo_{nro_comprobante}.pdf",
        "enlace_xml": f"https://api.tu-proveedor.com/xml/demo_{nro_comprobante}.xml"
    }

@app.route('/caja/cobrar/<int:orden_id>', methods=['GET', 'POST'])
@roles_required(['admin', 'cajero'])
def cobrar_orden(orden_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    usuario_id = session.get('user_id') or session.get('id')
    
    if request.method == 'POST':
        try:
            # 👉 AQUÍ ESTÁ LA SOLUCIÓN DEFINITIVA AL ERROR "Unread result found": buffered=True
            cursor = conn.cursor(dictionary=True, buffered=True)
            
            # Verificar caja abierta para registrar el ingreso de dinero
            cursor.execute("SELECT id FROM caja_sesiones WHERE usuario_id = %s AND estado = 'abierta'", (usuario_id,))
            sesion_caja = cursor.fetchone()
            
            if not sesion_caja:
                flash("Error: No tienes una caja abierta activa.", "danger")
                return redirect(url_for('caja_index'))

            tipo_comprobante = request.form['tipo_comprobante']
            metodo_pago = request.form['metodo_pago']
            monto_total = float(request.form['monto_total'])
            serie = EMPRESA_DATOS['serie_boleta'] if tipo_comprobante == 'boleta' else EMPRESA_DATOS['serie_factura']
            
            # Generar Nro Comprobante
            cursor.execute("SELECT MAX(CAST(SUBSTRING_INDEX(comprobante, '-', -1) AS UNSIGNED)) as max_num FROM pagos WHERE comprobante LIKE %s", (f"{serie}-%",))
            ultimo_numero = cursor.fetchone()['max_num']
            nuevo_numero = (ultimo_numero + 1) if ultimo_numero else 1
            nro_comprobante = f"{serie}-{nuevo_numero:08d}"
            
            # 1. Insertar Factura/Boleta en BD local (Inicialmente como 'procesando')
            cursor.execute("""
                INSERT INTO pagos (orden_id, monto_total, tipo_comprobante, metodo_pago, comprobante, estado_sunat) 
                VALUES (%s, %s, %s, %s, %s, 'procesando')
            """, (orden_id, monto_total, tipo_comprobante, metodo_pago, nro_comprobante))
            pago_id = cursor.lastrowid
            
            # 2. Insertar Dinero en la Caja
            concepto = f"Cobro OT-{orden_id:04d} / {nro_comprobante}"
            cursor.execute("""
                INSERT INTO caja_movimientos (caja_sesion_id, tipo_movimiento, monto, metodo_pago, concepto, pago_id)
                VALUES (%s, 'ingreso', %s, %s, %s, %s)
            """, (sesion_caja['id'], monto_total, metodo_pago, concepto, pago_id))
            
            # 3. Marcar orden como entregada
            cursor.execute("UPDATE ordenes_servicio SET estado = 'entregado' WHERE id = %s", (orden_id,))
            
            # =========================================================================
            # INTEGRACIÓN SUNAT (SIMULACIÓN PARA PRUEBAS)
            # =========================================================================
            # Recopilamos datos del cliente para la factura
            cursor.execute("""
                SELECT p.nombre_completo, p.numero_documento as dni_ruc, p.direccion, p.email 
                FROM ordenes_servicio os
                JOIN clientes c ON os.vehiculo_id = (SELECT id FROM vehiculos WHERE id = os.vehiculo_id LIMIT 1)
                JOIN personas p ON c.persona_id = p.id
                WHERE os.id = %s
            """, (orden_id,))
            datos_cliente = cursor.fetchone()
            
            # Recopilamos repuestos vendidos
            cursor.execute("""
                SELECT p.nombre, p.codigo, do.cantidad, do.precio_unitario, p.tipo 
                FROM detalles_orden do 
                JOIN productos p ON do.producto_id = p.id 
                WHERE do.orden_id = %s
            """, (orden_id,))
            detalles_venta = cursor.fetchall()
            
            # Cálculos de impuestos
            totales = {
                "subtotal": round(monto_total / 1.18, 2),
                "igv": round(monto_total - (monto_total / 1.18), 2),
                "total": round(monto_total, 2)
            }
            
            # ¡Llamamos al Simulador!
            try:
                # Nos aseguramos de que el método enviar_comprobante_sunat exista (lo pusimos en un paso anterior)
                respuesta_sunat = enviar_comprobante_sunat(pago_id, nro_comprobante, tipo_comprobante, datos_cliente, detalles_venta, totales)
                
                if respuesta_sunat['exito']:
                    cursor.execute("""
                        UPDATE pagos SET estado_sunat = %s, sunat_hash = %s 
                        WHERE id = %s
                    """, (respuesta_sunat['estado_sunat'], respuesta_sunat['sunat_hash'], pago_id))
                    
                    flash(f"💰 ¡Cobro registrado! La {tipo_comprobante.upper()} {nro_comprobante} fue generada exitosamente.", "success")
                else:
                    flash("⚠️ Cobro guardado en caja, pero falló el envío al facturador.", "warning")
            except NameError:
                # Si por alguna razón olvidaste pegar la función enviar_comprobante_sunat, la venta se guarda igual
                cursor.execute("UPDATE pagos SET estado_sunat = 'aceptado' WHERE id = %s", (pago_id,))
                flash(f"💰 ¡Cobro registrado con éxito en la caja! Comprobante: {nro_comprobante}", "success")

            conn.commit()
            
            # --- AUDITORÍA ---
            registrar_auditoria("Cobró orden y emitió comprobante", "CAJA", {"orden_id": orden_id, "comprobante": nro_comprobante, "monto": monto_total})
            
            return redirect(url_for('caja_index'))
            
        except Error as e:
            conn.rollback()
            flash(f"Error al cobrar: {e}", "danger")
        finally:
            if conn and conn.is_connected(): 
                cursor.close()
                conn.close()
        return redirect(url_for('caja_index'))

    # Mostrar la vista de cobro (GET)
    # 👉 AQUÍ TAMBIÉN ESTÁ LA SOLUCIÓN AL ERROR ROJO: buffered=True
    cursor = conn.cursor(dictionary=True, buffered=True)
    
    cursor.execute("""
        SELECT os.id, p.nombre_completo as cliente, p.numero_documento as dni_ruc, 
               v.placa, v.marca, v.modelo, os.fecha_ingreso 
        FROM ordenes_servicio os 
        JOIN vehiculos v ON os.vehiculo_id = v.id 
        JOIN clientes c ON v.cliente_id = c.id 
        JOIN personas p ON c.persona_id = p.id
        WHERE os.id = %s
    """, (orden_id,))
    orden = cursor.fetchone()
    
    cursor.execute("""
        SELECT p.nombre, p.codigo, do.cantidad, do.precio_unitario, (do.cantidad * do.precio_unitario) as subtotal 
        FROM detalles_orden do 
        JOIN productos p ON do.producto_id = p.id 
        WHERE do.orden_id = %s
    """, (orden_id,))
    detalles = cursor.fetchall()
    conn.close()
    
    total = sum(d['subtotal'] for d in detalles if d['subtotal'] is not None)
    subtotal_base = float(total) / 1.18
    igv = float(total) - subtotal_base
    trx_auto = f"TRX-{datetime.now().strftime('%Y%m%d')}-{orden_id:04d}"
    
    return render_template('caja/cobrar.html', orden=orden, detalles=detalles, total=total, subtotal=subtotal_base, igv=igv, trx_auto=trx_auto)


@app.route('/caja/cierre', methods=['GET', 'POST'])
@roles_required(['admin', 'cajero'])
def cierre_caja():
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))

    usuario_id = session.get('user_id') or session.get('id')
    cursor = conn.cursor(dictionary=True)

    # 1. Buscar caja abierta
    cursor.execute("SELECT * FROM caja_sesiones WHERE usuario_id = %s AND estado = 'abierta'", (usuario_id,))
    caja_abierta = cursor.fetchone()

    if not caja_abierta:
        flash("No tienes ninguna sesión de caja abierta para cerrar.", "warning")
        cursor.close()
        conn.close()
        return redirect(url_for('caja_index'))

    # 2. Calcular Efectivo Físico (Lo que audita)
    cursor.execute("""
        SELECT COALESCE(SUM(monto), 0) as total_ingresos 
        FROM caja_movimientos 
        WHERE caja_sesion_id = %s AND tipo_movimiento = 'ingreso' AND metodo_pago = 'efectivo'
    """, (caja_abierta['id'],))
    ingresos_efectivo = cursor.fetchone()['total_ingresos']

    # 2.1 Calcular Ingresos Digitales (Solo informativo)
    cursor.execute("""
        SELECT COALESCE(SUM(monto), 0) as total_digital 
        FROM caja_movimientos 
        WHERE caja_sesion_id = %s AND tipo_movimiento = 'ingreso' AND metodo_pago != 'efectivo'
    """, (caja_abierta['id'],))
    ingresos_digitales = cursor.fetchone()['total_digital']

    cursor.execute("""
        SELECT COALESCE(SUM(monto), 0) as total_egresos 
        FROM caja_movimientos 
        WHERE caja_sesion_id = %s AND tipo_movimiento = 'egreso' AND metodo_pago = 'efectivo'
    """, (caja_abierta['id'],))
    egresos_efectivo = cursor.fetchone()['total_egresos']

    # Matemática del Arqueo (Solo EFECTIVO)
    monto_calculado = float(caja_abierta['monto_inicial']) + float(ingresos_efectivo) - float(egresos_efectivo)

    # 3. Procesar Cierre
    if request.method == 'POST':
        try:
            monto_fisico = float(request.form['monto_fisico'].replace(',', '')) # Limpiamos formato por si acaso
            observaciones_cierre = request.form.get('observaciones', '')
            descuadre = monto_fisico - monto_calculado

            cursor.execute("""
                UPDATE caja_sesiones 
                SET fecha_cierre = NOW(), 
                    monto_calculado = %s, 
                    monto_fisico = %s, 
                    descuadre = %s, 
                    estado = 'cerrada', 
                    observaciones = CONCAT(COALESCE(observaciones, ''), ' | Cierre: ', %s)
                WHERE id = %s
            """, (monto_calculado, monto_fisico, descuadre, observaciones_cierre, caja_abierta['id']))
            conn.commit()

            # --- AUDITORÍA ---
            registrar_auditoria("Cerró la caja registradora", "CAJA", {"monto_fisico": monto_fisico, "descuadre": descuadre})

            return redirect(url_for('caja_index'))

        except Error as e:
            flash(f"Error al cerrar la caja: {e}", "danger")
        finally:
            cursor.close()
            conn.close()

    cursor.close()
    conn.close()
    # Pasamos las variables a la vista
    return render_template('caja/cierre_caja.html', caja=caja_abierta, ingresos_digitales=ingresos_digitales)

@app.route('/caja/resolver/<int:caja_id>', methods=['POST'])
@roles_required(['admin']) # MUY IMPORTANTE: Solo el administrador puede hacer esto
def resolver_descuadre(caja_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    try:
        cursor = conn.cursor()
        # Actualizamos el descuadre a 0 y dejamos una marca imborrable en auditoría
        cursor.execute("""
            UPDATE caja_sesiones 
            SET descuadre = 0, 
                observaciones = CONCAT(COALESCE(observaciones, ''), ' | [AUDITORÍA]: Faltante repuesto por el empleado y validado por Administración.')
            WHERE id = %s
        """, (caja_id,))
        conn.commit()
        
        flash("Faltante de caja justificado y repuesto correctamente. El turno ha sido liberado de alertas.", "success")
        
    except Error as e:
        flash(f"Error al validar el descuadre: {e}", "danger")
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
            
    # Lo regresamos a la caja, que ahora le mostrará la pantalla verde
    return redirect(url_for('caja_index'))

# =================================================================================
# FUNCIÓN DE ANULACIÓN 
# =================================================================================
@app.route('/caja/anular/<int:pago_id>', methods=['POST'])
@roles_required(['admin']) # GUARDIA DE SEGURIDAD: Solo el admin puede hacer esto
def anular_pago(pago_id):
    conn = get_db_connection()
    if not conn: return redirect(url_for('dashboard'))
    
    try:
        cursor = conn.cursor(dictionary=True)
        
        # 1. Traer los datos del pago que queremos anular
        cursor.execute("SELECT orden_id, monto_total, metodo_pago, comprobante FROM pagos WHERE id = %s", (pago_id,))
        pago = cursor.fetchone()
        
        if not pago:
            flash("Error: El comprobante no existe.", "danger")
            return redirect(url_for('caja_index'))
            
        # 2. Buscar la caja ABIERTA del administrador para sacarle el dinero
        usuario_id = session.get('user_id') or session.get('id')
        cursor.execute("SELECT id FROM caja_sesiones WHERE usuario_id = %s AND estado = 'abierta'", (usuario_id,))
        sesion_caja = cursor.fetchone()
        
        if not sesion_caja:
            flash("Para anular un comprobante, necesitas tener tu caja abierta (fondo inicial) para registrar la salida del dinero.", "warning")
            return redirect(url_for('caja_index'))

        # 3. MARCAR COMO ANULADO (Auditoría)
        cursor.execute("UPDATE pagos SET estado = 'anulado' WHERE id = %s", (pago_id,))
        
        # 4. REGRESAR LA ORDEN A PENDIENTE
        cursor.execute("UPDATE ordenes_servicio SET estado = 'finalizado' WHERE id = %s", (pago['orden_id'],))
        
        # 5. RETIRAR EL DINERO DE LA GAVETA (Egreso)
        concepto = f"EXTORNO AUDITORÍA: Anulación de {pago['comprobante']}"
        cursor.execute("""
            INSERT INTO caja_movimientos (caja_sesion_id, tipo_movimiento, monto, metodo_pago, concepto, pago_id)
            VALUES (%s, 'egreso', %s, %s, %s, %s)
        """, (sesion_caja['id'], pago['monto_total'], pago['metodo_pago'], concepto, pago_id))
        
        conn.commit()
        flash(f"¡Auditoría: Comprobante {pago['comprobante']} anulado! El dinero fue descontado y la orden volvió a pendientes.", "success")
        
    except Error as e:
        flash(f"Error al anular: {e}", "danger")
    finally:
        if conn.is_connected():
            cursor.close()
            conn.close()
            
    return redirect(url_for('caja_index'))

# ==============================================================================
# API AJAX: VER DETALLES DE UNA VENTA/COBRO DESDE LA CAJA
# ==============================================================================
@app.route('/api/ventas/<int:pago_id>/detalles')
@roles_required(['admin', 'cajero'])
def api_detalles_venta(pago_id):
    conn = get_db_connection()
    if not conn:
        return jsonify({"error": "Error de conexión a la BD"}), 500
    
    try:
        cursor = conn.cursor(dictionary=True)
        # 1. Traer datos financieros del pago
        cursor.execute("""
            SELECT pg.monto_total, pg.comprobante, pg.metodo_pago, pg.estado, os.id as orden_id
            FROM pagos pg
            JOIN ordenes_servicio os ON pg.orden_id = os.id
            WHERE pg.id = %s
        """, (pago_id,))
        pago = cursor.fetchone()
        
        if not pago:
            return jsonify({"error": "Pago no encontrado"}), 404
            
        # 2. Traer los repuestos y servicios de esa orden
        cursor.execute("""
            SELECT p.codigo, p.nombre, do.cantidad, do.precio_unitario as precio, 
                   (do.cantidad * do.precio_unitario) as subtotal
            FROM detalles_orden do
            JOIN productos p ON do.producto_id = p.id
            WHERE do.orden_id = %s
        """, (pago['orden_id'],))
        detalles = cursor.fetchall()
        
        total = float(pago['monto_total'])
        subtotal_base = total / 1.18
        igv = total - subtotal_base

        return jsonify({
            "resumen": {
                "comprobante": pago['comprobante'],
                "metodo_pago": pago['metodo_pago'],
                "estado": pago['estado'],
                "subtotal": subtotal_base,
                "igv": igv,
                "total": total
            },
            "items": detalles
        })
    except Error as e:
        return jsonify({"error": str(e)}), 500
    finally:
        cursor.close()
        conn.close()

# ==============================================================================
# REABRIR CAJA (SALVAVIDAS PARA CIERRES ACCIDENTALES)
# ==============================================================================
@app.route('/caja/reabrir/<int:sesion_id>', methods=['POST'])
@roles_required(['admin'])  # <-- ¡CRÍTICO! Solo el admin puede hacer esto
def reabrir_caja(sesion_id):
    conn = get_db_connection()
    if not conn:
        flash("Error de conexión a la base de datos.", "danger")
        return redirect(url_for('caja_index'))

    cursor = conn.cursor(dictionary=True)
    try:
        # 1. Verificar el estado actual de esa caja
        cursor.execute("SELECT estado, usuario_id FROM caja_sesiones WHERE id = %s", (sesion_id,))
        caja = cursor.fetchone()

        if not caja:
            flash("La sesión de caja no existe.", "danger")
            return redirect(url_for('caja_index'))

        if caja['estado'] == 'abierta':
            flash("Esta caja ya se encuentra abierta.", "info")
            return redirect(url_for('caja_index'))

        # 2. La Magia: Revivir la caja (Cambiar a abierta y borrar la fecha de cierre)
        cursor.execute("""
            UPDATE caja_sesiones 
            SET estado = 'abierta', 
                fecha_cierre = NULL 
            WHERE id = %s
        """, (sesion_id,))

        conn.commit()

        # 3. Dejar rastro en la Auditoría (Prueba de Caja Blanca)
        registrar_auditoria(
            "Reabrió caja cerrada por error", 
            "CAJA", 
            {"sesion_id": sesion_id, "cajero_id": caja['usuario_id']}
        )

        flash(f"✅ ¡Salvavidas activado! La caja #{sesion_id} ha sido reabierta. El cajero puede seguir cobrando con normalidad.", "success")
        
    except Error as e:
        conn.rollback()
        flash(f"Error al intentar reabrir la caja: {e.msg}", "danger")
    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('caja_index'))

# =================================================================================
# GENERADOR DE COMPROBANTES PDF A4 (CON DATOS SUNAT EXACTOS)
# =================================================================================
class PDF(FPDF):
    def __init__(self, tipo_comprobante, numero_comprobante, fecha):
        super().__init__(orientation='P', unit='mm', format='A4')
        self.tipo_comprobante = tipo_comprobante.upper()
        self.numero_comprobante = numero_comprobante
        self.fecha = fecha
        
        # DATOS REALES DEL SISTEMA
        try:
            self.empresa_nombre = EMPRESA_DATOS['nombre'].encode('latin-1', 'replace').decode('latin-1')
            self.empresa_ruc = EMPRESA_DATOS['ruc']
            self.empresa_dir = EMPRESA_DATOS['direccion'].encode('latin-1', 'replace').decode('latin-1')
            self.empresa_tel = EMPRESA_DATOS['telefono']
            self.empresa_email = EMPRESA_DATOS.get('email', '')
        except:
            self.empresa_nombre = "MULTISERVICIOS VIRGEN DE CHAPI"
            self.empresa_ruc = "10806325631"
            self.empresa_dir = "Carretera a Puquio KM. 2.5, Pj. Buena Fe"
            self.empresa_tel = "966633037 / 056-523915"
            self.empresa_email = "moralescaritas_@hotmail.com"
            
        self.set_auto_page_break(True, 35)

    def header(self):
        # 1. LOGO 
        logo_path = 'static/logo_chapi.png'
        if os.path.exists(logo_path):
            self.image(logo_path, 10, 10, 35)
            
        # 2. DATOS EMPRESA 
        self.set_xy(45, 12)
        self.set_font('Arial', 'B', 12) 
        self.cell(90, 6, self.empresa_nombre, 0, 1, 'C')
        
        self.set_font('Arial', '', 8)
        self.set_x(45)
        self.cell(90, 4, self.empresa_dir, 0, 1, 'C')
        self.set_x(45)
        self.cell(90, 4, f"Telf: {self.empresa_tel}", 0, 1, 'C')
        self.set_x(45)
        if self.empresa_email:
            self.cell(90, 4, f"Email: {self.empresa_email}", 0, 1, 'C')
        
        # 3. RECUADRO SUNAT
        self.set_xy(135, 10)
        self.rect(135, 10, 65, 25, 'D') 
        
        self.set_xy(135, 12)
        self.set_font('Arial', 'B', 11)
        self.cell(65, 6, f"R.U.C. {self.empresa_ruc}", 0, 1, 'C')
        
        self.set_x(135)
        self.set_font('Arial', 'B', 10) 
        self.cell(65, 6, f"{self.tipo_comprobante}", 0, 1, 'C')
        
        self.set_x(135)
        self.set_font('Arial', 'B', 11)
        self.cell(65, 6, f"Nro: {self.numero_comprobante}", 0, 1, 'C')
        self.ln(15)

    def datos_cliente(self, cliente, ruc_dni, direccion, fecha_emision, orden_id):
        self.set_y(48)
        self.rect(10, 48, 190, 22, 'D')
        
        self.set_font('Arial', 'B', 8)
        self.set_xy(12, 51)
        self.cell(22, 5, 'CLIENTE:', 0, 0)
        self.set_font('Arial', '', 8)
        self.cell(100, 5, cliente.encode('latin-1', 'replace').decode('latin-1'), 0, 1)
        
        self.set_xy(12, 56)
        self.set_font('Arial', 'B', 8)
        self.cell(22, 5, 'DIRECCION:', 0, 0)
        self.set_font('Arial', '', 8)
        self.cell(100, 5, direccion.encode('latin-1', 'replace').decode('latin-1') if direccion else '---', 0, 1)
        
        self.set_xy(12, 61)
        self.set_font('Arial', 'B', 8)
        self.cell(22, 5, 'DNI / RUC:', 0, 0)
        self.set_font('Arial', '', 8)
        self.cell(100, 5, ruc_dni, 0, 1)
        
        self.set_y(73)
        self.set_font('Arial', 'B', 7)
        self.cell(40, 5, 'FECHA EMISION', 1, 0, 'C')
        self.cell(40, 5, 'FEC. VENCIMIENTO', 1, 0, 'C')
        self.cell(60, 5, 'GUIA / ORDEN', 1, 0, 'C')
        self.cell(50, 5, 'COND. DE PAGO', 1, 1, 'C')
        
        # AQUÍ ESTÁ LA MAGIA: Ingresa la Orden (OT-XXXX) y dice CONTADO
        self.set_font('Arial', '', 7)
        self.cell(40, 5, fecha_emision, 1, 0, 'C')
        self.cell(40, 5, fecha_emision, 1, 0, 'C')
        self.cell(60, 5, f"OT-{orden_id:04d}", 1, 0, 'C')
        self.cell(50, 5, "CONTADO", 1, 1, 'C')
        self.ln(3)

    def tabla_items(self, detalles, subtotal_base, igv, total, total_en_palabras, qr_path):
        self.set_font('Arial', 'B', 8)
        self.cell(20, 6, 'CANTIDAD', 1, 0, 'C')
        self.cell(15, 6, 'U.M', 1, 0, 'C')
        self.cell(105, 6, 'DESCRIPCION', 1, 0, 'C')
        self.cell(25, 6, 'PRECIO UNIT.', 1, 0, 'C')
        self.cell(25, 6, 'IMPORTE', 1, 1, 'C')
        
        self.set_font('Arial', '', 8)
        start_y = self.get_y()
        for item in detalles:
            self.cell(20, 5, str(item['cantidad']), 'LR', 0, 'C')
            self.cell(15, 5, 'NIU', 'LR', 0, 'C')
            self.cell(105, 5, item['nombre'].encode('latin-1', 'replace').decode('latin-1'), 'LR', 0, 'L')
            self.cell(25, 5, f"{item['precio_unitario']:.2f}", 'LR', 0, 'R')
            self.cell(25, 5, f"{item['subtotal']:.2f}", 'LR', 1, 'R')
            
        end_y = self.get_y()
        altura_minima = 80 
        relleno = altura_minima - (end_y - start_y) if (end_y - start_y) < altura_minima else 5
            
        self.rect(10, start_y, 20, end_y - start_y + relleno)
        self.rect(30, start_y, 15, end_y - start_y + relleno)
        self.rect(45, start_y, 105, end_y - start_y + relleno)
        self.rect(150, start_y, 25, end_y - start_y + relleno)
        self.rect(175, start_y, 25, end_y - start_y + relleno)
        
        self.set_y(end_y + relleno + 2)
        
        self.set_font('Arial', 'B', 8)
        self.cell(0, 5, total_en_palabras, 0, 1, 'L')
        
        y_totales = self.get_y()
        if os.path.exists(qr_path):
            self.image(qr_path, 10, y_totales, 30, 30)
            os.remove(qr_path)
            
        self.set_xy(140, y_totales + 5)
        self.cell(35, 6, "OP. GRAVADA (S/)", 0, 0, 'R')
        self.cell(25, 6, f"{subtotal_base:.2f}", 0, 1, 'R')
        self.set_x(140)
        self.cell(35, 6, "TOTAL IGV (S/)", 0, 0, 'R')
        self.cell(25, 6, f"{igv:.2f}", 'B', 1, 'R')
        self.set_x(140)
        self.cell(35, 6, "IMPORTE TOTAL (S/)", 0, 0, 'R')
        self.cell(25, 6, f"{total:.2f}", 0, 1, 'R')
        
        self.ln(15)
        self.set_font('Arial', '', 7)
        texto_pie = f"Representacion Impresa de la {self.tipo_comprobante}, Autorizado por Resolucion 034-005-0007241/SUNAT."
        self.cell(0, 4, texto_pie, 0, 1, 'C')


@app.route('/caja/comprobante/<int:pago_id>')
def imprimir_comprobante(pago_id):
    conn = get_db_connection()
    if not conn: return "Error de BD", 500
    
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT pg.*, p.nombre_completo as nombre_razon_social, p.numero_documento as dni_ruc, p.direccion,
               os.id as orden_id
        FROM pagos pg 
        JOIN ordenes_servicio os ON pg.orden_id = os.id
        JOIN vehiculos v ON os.vehiculo_id = v.id
        JOIN clientes c ON v.cliente_id = c.id
        JOIN personas p ON c.persona_id = p.id
        WHERE pg.id = %s
    """, (pago_id,))
    pago = cursor.fetchone()
    
    if not pago: return redirect(url_for('caja_index'))
        
    cursor.execute("""
        SELECT p.nombre, p.codigo, do.cantidad, do.precio_unitario, (do.cantidad * do.precio_unitario) as subtotal 
        FROM detalles_orden do 
        JOIN productos p ON do.producto_id = p.id 
        WHERE do.orden_id = %s
    """, (pago['orden_id'],))
    detalles = cursor.fetchall()
    conn.close()
    
    total = float(pago['monto_total'])
    subtotal_base = total / 1.18
    igv = total - subtotal_base
    total_int = int(total)
    total_dec = int(round((total - total_int) * 100))
    total_en_palabras = f"SON: {num2words(total_int, lang='es').upper()} Y {total_dec:02d}/100 SOLES"
    
    ruc_empresa = EMPRESA_DATOS.get('ruc', '10806325631')
    qr_str = f"{ruc_empresa}|{pago['tipo_comprobante']}|{pago['comprobante']}|{igv:.2f}|{total:.2f}|{pago['fecha_pago'].strftime('%Y-%m-%d')}|{pago['dni_ruc']}"
    qr_img = qrcode.make(qr_str)
    qr_path = f"temp_qr_{pago_id}.png"
    qr_img.save(qr_path)

    estado_comp = f"{pago['tipo_comprobante']} DE VENTA ELECTRONICA" if pago['tipo_comprobante'] == 'boleta' else f"{pago['tipo_comprobante']} ELECTRONICA"
    if pago['estado'] == 'anulado': estado_comp += " (ANULADA)"

    pdf = PDF(estado_comp, pago['comprobante'], pago['fecha_pago'].strftime('%d/%m/%Y'))
    pdf.add_page()
    
    # Enviamos el orden_id en lugar de la condición de pago
    pdf.datos_cliente(pago['nombre_razon_social'], pago['dni_ruc'], pago['direccion'], pago['fecha_pago'].strftime('%d/%m/%Y'), pago['orden_id'])
    
    pdf.tabla_items(detalles, subtotal_base, igv, total, total_en_palabras, qr_path)
    
    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers.set('Content-Disposition', 'attachment', filename=f"{pago['comprobante']}.pdf")
    response.headers.set('Content-Type', 'application/pdf')
    return response


# =================================================================================
# GENERADOR DE TICKET TÉRMICO (ACTUALIZADO CON NÚMERO DE ORDEN)
# =================================================================================
@app.route('/caja/ticket/<int:pago_id>')
def imprimir_ticket(pago_id):
    conn = get_db_connection()
    if not conn: return "Error de BD", 500
    
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT pg.*, p.nombre_completo as nombre_razon_social, p.numero_documento as dni_ruc, p.direccion,
               os.id as orden_id
        FROM pagos pg 
        JOIN ordenes_servicio os ON pg.orden_id = os.id
        JOIN vehiculos v ON os.vehiculo_id = v.id
        JOIN clientes c ON v.cliente_id = c.id
        JOIN personas p ON c.persona_id = p.id
        WHERE pg.id = %s
    """, (pago_id,))
    pago = cursor.fetchone()
    
    if not pago: return redirect(url_for('caja_index'))
        
    cursor.execute("""
        SELECT p.nombre, p.codigo, do.cantidad, do.precio_unitario, (do.cantidad * do.precio_unitario) as subtotal 
        FROM detalles_orden do 
        JOIN productos p ON do.producto_id = p.id 
        WHERE do.orden_id = %s
    """, (pago['orden_id'],))
    detalles = cursor.fetchall()
    conn.close()
    
    try:
        empresa_nombre = EMPRESA_DATOS['nombre'].encode('latin-1', 'replace').decode('latin-1')
        empresa_ruc = EMPRESA_DATOS['ruc']
        empresa_dir = EMPRESA_DATOS['direccion'].encode('latin-1', 'replace').decode('latin-1')
        empresa_tel = EMPRESA_DATOS['telefono']
    except:
        empresa_nombre = "MULTISERVICIOS VIRGEN DE CHAPI"
        empresa_ruc = "10806325631"
        empresa_dir = "Carretera a Puquio KM. 2.5"
        empresa_tel = "966633037"

    total = float(pago['monto_total'])
    subtotal_base = total / 1.18
    igv = total - subtotal_base
    total_int = int(total)
    total_dec = int(round((total - total_int) * 100))
    total_en_palabras = f"SON: {num2words(total_int, lang='es').upper()} Y {total_dec:02d}/100 SOLES"
    
    pdf = FPDF(orientation='P', unit='mm', format=(80, 250))
    pdf.set_margins(4, 5, 4)
    pdf.add_page()
    
    logo_path = 'static/logo_chapi.png'
    if os.path.exists(logo_path):
        pdf.image(logo_path, 20, 5, 40)
        pdf.set_y(48)
    else:
        pdf.set_y(5)
    
    pdf.set_font('Arial', 'B', 10)
    pdf.cell(0, 4, empresa_nombre, 0, 1, 'C')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(0, 5, f"RUC: {empresa_ruc}", 0, 1, 'C')
    
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 4, empresa_dir, 0, 1, 'C')
    pdf.cell(0, 4, f"Telf: {empresa_tel}", 0, 1, 'C')
    pdf.ln(4)
    
    pdf.set_font('Arial', 'B', 10)
    estado_comp = f"{pago['tipo_comprobante'].upper()} DE VENTA ELECTRÓNICA" if pago['tipo_comprobante'] == 'boleta' else f"{pago['tipo_comprobante'].upper()} ELECTRÓNICA"
    if pago['estado'] == 'anulado': estado_comp += " (ANULADA)"
    
    pdf.cell(0, 5, estado_comp, 0, 1, 'C')
    pdf.cell(0, 5, pago['comprobante'], 0, 1, 'C')
    pdf.ln(3)
    
    pdf.set_font('Arial', '', 9)
    pdf.cell(0, 5, f"{pago['nombre_razon_social'][:35].encode('latin-1', 'replace').decode('latin-1')}", 0, 1, 'C')
    pdf.cell(0, 3, "---", 0, 1, 'C')
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(0, 5, f"DNI/RUC {pago['dni_ruc']}", 0, 1, 'C')
    pdf.ln(2)
    
    # ------------- AQUÍ AGREGAMOS LA ORDEN AL TICKET -------------
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(15, 5, "FECHA:", 0, 0, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.cell(20, 5, pago['fecha_pago'].strftime('%d/%m/%Y'), 0, 0, 'L')
    
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(15, 5, "HORA:", 0, 0, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 5, pago['fecha_pago'].strftime('%I:%M %p'), 0, 1, 'L')

    pdf.set_font('Arial', 'B', 8)
    pdf.cell(15, 5, "ORDEN:", 0, 0, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 5, f"OT-{pago['orden_id']:04d}", 0, 1, 'L')
    # -------------------------------------------------------------
    
    pdf.line(4, pdf.get_y(), 76, pdf.get_y())
    pdf.ln(1)
    
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(8, 5, "Cant", 0, 0, 'L')
    pdf.cell(8, 5, "U.M", 0, 0, 'L')
    pdf.cell(20, 5, "COD", 0, 0, 'L')
    pdf.cell(16, 5, "PRECIO", 0, 0, 'R')
    pdf.cell(20, 5, "TOTAL", 0, 1, 'R')
    pdf.cell(0, 5, "DESCRIPCION", 0, 1, 'L')
    
    pdf.line(4, pdf.get_y(), 76, pdf.get_y())
    pdf.ln(1)
    
    pdf.set_font('Arial', '', 8)
    for item in detalles:
        pdf.cell(8, 5, str(item['cantidad']), 0, 0, 'L')
        pdf.cell(8, 5, 'NIU', 0, 0, 'L')
        codigo = item.get('codigo', 'S/N')
        pdf.cell(20, 5, str(codigo)[:10], 0, 0, 'L')
        pdf.cell(16, 5, f"{item['precio_unitario']:.2f}", 0, 0, 'R')
        pdf.cell(20, 5, f"{item['subtotal']:.2f}", 0, 1, 'R')
        
        desc_corta = item['nombre'][:38].encode('latin-1', 'replace').decode('latin-1')
        pdf.cell(0, 5, desc_corta, 0, 1, 'L')
        
    pdf.line(4, pdf.get_y(), 76, pdf.get_y())
    pdf.ln(2)
    
    pdf.set_font('Arial', 'B', 9)
    pdf.cell(40, 5, "TOTAL GRAVADO", 0, 0, 'L')
    pdf.cell(10, 5, "(S/)", 0, 0, 'C')
    pdf.cell(22, 5, f"{subtotal_base:.2f}", 0, 1, 'R')
    
    pdf.cell(40, 5, "I.G.V", 0, 0, 'L')
    pdf.cell(10, 5, "(S/)", 0, 0, 'C')
    pdf.cell(22, 5, f"{igv:.2f}", 0, 1, 'R')
    
    pdf.set_font('Arial', 'B', 12) 
    pdf.cell(40, 6, "TOTAL", 0, 0, 'L')
    pdf.cell(10, 6, "(S/)", 0, 0, 'C')
    pdf.cell(22, 6, f"{total:.2f}", 0, 1, 'R')
    pdf.ln(2)
    
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(10, 5, "SON: ", 0, 0, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 5, total_en_palabras, 0, 1, 'L')
    
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(30, 5, "FORMA DE PAGO: ", 0, 0, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 5, pago['metodo_pago'].upper(), 0, 1, 'L')
    
    pdf.set_font('Arial', 'B', 8)
    pdf.cell(25, 5, "COND.VENTA: ", 0, 0, 'L')
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 5, "CONTADO", 0, 1, 'L')
    
    qr_str = f"{empresa_ruc}|{pago['tipo_comprobante']}|{pago['comprobante']}|{igv:.2f}|{total:.2f}|{pago['fecha_pago'].strftime('%Y-%m-%d')}|{pago['dni_ruc']}"
    qr_img = qrcode.make(qr_str)
    qr_path = f"temp_qr_tkt_{pago_id}.png"
    qr_img.save(qr_path)
    
    pdf.ln(4)
    if os.path.exists(qr_path):
        pdf.image(qr_path, 25, pdf.get_y(), 30, 30)
        os.remove(qr_path)
        pdf.ln(32)
        
    tipo_legal = "BOLETA DE VENTA" if pago['tipo_comprobante'] == 'boleta' else "FACTURA"
    pdf.set_font('Arial', '', 8)
    pdf.cell(0, 4, f"Representacion Impresa de la {tipo_legal}", 0, 1, 'C')
    pdf.cell(0, 4, "ELECTRONICA", 0, 1, 'C')
    pdf.cell(0, 4, "Puede consultar en el sistema", 0, 1, 'C')
    pdf.cell(0, 4, "Autorizado mediante Resolucion de SUNAT", 0, 1, 'C')
    
    response = make_response(pdf.output(dest='S').encode('latin-1'))
    response.headers.set('Content-Disposition', 'attachment', filename=f"TKT_{pago['comprobante']}.pdf")
    response.headers.set('Content-Type', 'application/pdf')
    return response

# ================== MÓDULO GESTIÓN DE USUARIOS (SOLO ADMIN) ==================

@app.route('/usuarios')
@roles_required(['admin'])
def usuarios_index():
    conn = get_db_connection()
    usuarios = []
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT u.id, p.nombre_completo as nombre, u.usuario, u.rol, p.email, u.activo 
            FROM usuarios u
            JOIN personas p ON u.persona_id = p.id
            ORDER BY p.nombre_completo
        """)
        usuarios = cursor.fetchall()
        conn.close()
    return render_template('usuarios/listar_usuarios.html', usuarios=usuarios)

@app.route('/usuarios/nuevo', methods=['GET', 'POST'])
@roles_required(['admin'])
def nuevo_usuario():
    if request.method == 'POST':
        try:
            conn = get_db_connection()
            cursor = conn.cursor()
            
            nombres = request.form.get('nombres', '').strip().upper()
            ap_paterno = request.form.get('ap_paterno', '').strip().upper()
            ap_materno = request.form.get('ap_materno', '').strip().upper()
            nombre_completo = f"{nombres} {ap_paterno} {ap_materno}".strip()

            cod_pais = request.form.get('cod_pais', '+51')
            num_tel = request.form.get('telefono', '').strip()
            telefono_completo = f"{cod_pais} {num_tel}"

            direccion = request.form.get('direccion', '').strip().upper()
            referencia = request.form.get('referencia', '').strip().upper()
            direccion_completa = f"{direccion} - REF: {referencia}" if referencia else direccion

            em_nombre = request.form.get('em_nombre', '').strip().upper()
            em_tel = request.form.get('em_telefono', '').strip()
            contacto_emergencia = f"{em_nombre} - {em_tel}" if em_nombre else ""

            # 1. Insertamos en PERSONAS
            cursor.execute("""
                INSERT INTO personas (tipo_documento, numero_documento, nombre_completo, telefono, email, direccion) 
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                request.form.get('tipo_documento', 'DNI'),
                request.form.get('numero_documento', ''),
                nombre_completo,
                telefono_completo,
                request.form.get('email', '').strip().lower(),
                direccion_completa
            ))
            persona_id = cursor.lastrowid 
            
            # 2. Procesamos fechas, sueldo y ubigeo
            fecha_ing = request.form.get('fecha_ingreso') or None
            fecha_nac = request.form.get('fecha_nacimiento') or None
            sueldo = request.form.get('sueldo_base') or None

            dep = request.form.get('departamento', '').strip().upper()
            prov = request.form.get('provincia', '').strip().upper()
            dist = request.form.get('distrito', '').strip().upper()
            ubigeo_completo = f"{dep} - {prov} - {dist}" if dep else ""

            nivel = request.form.get('nivel_cargo', '')
            area = request.form.get('especialidad_area', '')
            experiencia = request.form.get('experiencia', '0')
            cargo_completo = f"{nivel} en {area} ({experiencia} años exp.)" if nivel and area else "General"

            # --- SEGURIDAD: Encriptamos la clave ---
            clave_encriptada = generate_password_hash(request.form['password'])

            # 3. Insertamos en USUARIOS con el Hash
            cursor.execute("""
                INSERT INTO usuarios (
                    persona_id, usuario, password, rol, activo, 
                    especialidad, contacto_emergencia, fecha_ingreso,
                    lugar_procedencia, fecha_nacimiento, tipo_sangre, condiciones_medicas,
                    sueldo_base, cuenta_bancaria
                ) VALUES (%s, %s, %s, %s, 1, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                persona_id, 
                request.form['usuario'], 
                clave_encriptada, # AQUÍ GUARDAMOS LA CLAVE ENCRIPTADA
                request.form['rol'],
                cargo_completo,
                contacto_emergencia,
                fecha_ing,
                ubigeo_completo,  
                fecha_nac,
                request.form.get('tipo_sangre', ''),
                request.form.get('condiciones_medicas', '').strip().upper(),
                sueldo,
                request.form.get('cuenta_bancaria', '')
            ))
            
            conn.commit()
            # --- AUDITORÍA ---
            registrar_auditoria("Registró un nuevo empleado", "RRHH", {"usuario": request.form['usuario'], "rol": request.form['rol']})
            conn.close()
            flash("✅ Trabajador registrado exitosamente en la planilla.", "success")
            return redirect(url_for('usuarios_index'))
        except Exception as e:
            if hasattr(e, 'errno') and e.errno == 1062:
                flash("❌ Error: El DNI o el Nombre de Usuario ya existen en el sistema.", "danger")
            else:
                flash(f"Error en BD: {str(e)}", "danger")
                
    return render_template('usuarios/nuevo_usuario.html')

@app.route('/usuarios/editar/<int:usuario_id>', methods=['GET', 'POST'])
@roles_required(['admin'])
def editar_usuario(usuario_id):
    conn = get_db_connection()
    if request.method == 'POST':
        try:
            password_ingresada = request.form['password']
            cursor = conn.cursor()
            
            # 1. Buscar el persona_id
            cursor.execute("SELECT persona_id FROM usuarios WHERE id = %s", (usuario_id,))
            persona_id = cursor.fetchone()[0]
            
            # 2. Actualizar PERSONAS
            cursor.execute(
                "UPDATE personas SET nombre_completo=%s, email=%s WHERE id=%s",
                (request.form['nombre'], request.form.get('email', ''), persona_id)
            )
            
            # 3. Actualizar USUARIOS
            if password_ingresada:
                # Si el admin escribió una clave nueva, la ENCRIPTAMOS antes de guardarla
                nueva_clave_hash = generate_password_hash(password_ingresada)
                cursor.execute(
                    "UPDATE usuarios SET usuario=%s, rol=%s, activo=%s, password=%s WHERE id=%s",
                    (request.form['usuario'], request.form['rol'], request.form['activo'], nueva_clave_hash, usuario_id)
                )
            else:
                # Si dejó la clave en blanco, actualizamos el resto sin tocar la clave
                cursor.execute(
                    "UPDATE usuarios SET usuario=%s, rol=%s, activo=%s WHERE id=%s",
                    (request.form['usuario'], request.form['rol'], request.form['activo'], usuario_id)
                )
            
            conn.commit()
            conn.close()
            flash("✅ Permisos y datos de usuario actualizados.", "info")
            return redirect(url_for('usuarios_index'))
        except Exception as e:
            flash(f"Error en BD: {str(e)}", "danger")
            
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT u.id, p.nombre_completo as nombre, u.usuario, u.rol, p.email, u.activo 
        FROM usuarios u
        JOIN personas p ON u.persona_id = p.id
        WHERE u.id = %s
    """, (usuario_id,))
    usuario = cursor.fetchone()
    conn.close()
    
    return render_template('usuarios/editar_usuario.html', u=usuario)

# ==============================================================================
# MÓDULO DE PROVEEDORES (HERENCIA DE PERSONAS)
# ==============================================================================
@app.route('/proveedores/nuevo', methods=['GET', 'POST'])
@roles_required(['admin', 'cajero']) # Solo los administradores o caja registran proveedores
def nuevo_proveedor():
    if request.method == 'POST':
        # 1. Capturamos los datos base (Persona)
        tipo_doc = request.form.get('tipo_documento', 'RUC')
        numero_doc = request.form.get('numero_documento', '').strip()
        nombre = request.form.get('nombre_completo', '').strip()
        telefono = request.form.get('telefono', '').strip()
        email = request.form.get('email', '').strip()
        direccion = request.form.get('direccion', '').strip()
        departamento = request.form.get('departamento', '')
        provincia = request.form.get('provincia', '')
        distrito = request.form.get('distrito', '')
        
        # 2. Capturamos los datos específicos del Proveedor
        nombre_contacto = request.form.get('nombre_contacto', '').strip()
        sitio_web = request.form.get('sitio_web', '').strip()
        
        conn = get_db_connection()
        if conn:
            try:
                cursor = conn.cursor(dictionary=True)
                
                # A. ¿Esta persona/empresa ya existe en la base maestra?
                cursor.execute("SELECT id FROM personas WHERE numero_documento = %s", (numero_doc,))
                persona_existente = cursor.fetchone()
                
                if persona_existente:
                    persona_id = persona_existente['id']
                    # Validamos que no tenga el rol de proveedor duplicado
                    cursor.execute("SELECT id FROM proveedores WHERE persona_id = %s", (persona_id,))
                    if cursor.fetchone():
                        flash("❌ Esta empresa ya estaba registrada como proveedor.", "warning")
                        return redirect(url_for('nuevo_proveedor')) # Redirigimos aquí mismo por ahora
                else:
                    # B. Lo insertamos desde cero en Personas
                    cursor.execute("""
                        INSERT INTO personas (tipo_documento, numero_documento, nombre_completo, telefono, email, direccion, departamento, provincia, distrito) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """, (tipo_doc, numero_doc, nombre, telefono, email, direccion, departamento, provincia, distrito))
                    persona_id = cursor.lastrowid
                
                # C. Le asignamos la "Camiseta" de PROVEEDOR
                cursor.execute("""
                    INSERT INTO proveedores (persona_id, nombre_contacto, sitio_web) 
                    VALUES (%s, %s, %s)
                """, (persona_id, nombre_contacto, sitio_web))
                
                conn.commit()
                flash("✅ Proveedor registrado con éxito usando Arquitectura de Herencia.", "success")
                return redirect(url_for('nuevo_proveedor')) # En el futuro lo mandaremos al listado de proveedores
                
            except Error as e:
                flash(f"Error en la base de datos: {e.msg}", "danger")
            finally:
                cursor.close()
                conn.close()
                
    return render_template('proveedores/nuevo_proveedor.html')

# ==============================================================================
# API PARA CONSULTAR DOCUMENTOS (BÚSQUEDA INTELIGENTE FASE 1 Y 2)
# ==============================================================================
@app.route('/api/consultar_documento', methods=['POST'])
def consultar_documento():
    data = request.get_json()
    tipo_doc = data.get('tipo', 'DNI')
    numero = data.get('numero', '').strip()

    if not numero:
        return jsonify({'success': False, 'error': 'Número no proporcionado'})

    # -------------------------------------------------------
    # FASE 1: BÚSQUEDA LOCAL (EN TU BASE DE DATOS)
    # -------------------------------------------------------
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM personas WHERE numero_documento = %s", (numero,))
        persona_local = cursor.fetchone()
        cursor.close()
        conn.close()

        if persona_local:
            # ¡La persona ya existe en el sistema! (Ejemplo: Susy Rivera)
            return jsonify({
                'success': True,
                'origen': 'local',
                'nombre': persona_local['nombre_completo'],
                'telefono': persona_local.get('telefono', ''),
                'correo': persona_local.get('email', ''),
                'direccion': persona_local.get('direccion', ''),
                'departamento': persona_local.get('departamento', ''), # <-- ESTO FALTABA
                'provincia': persona_local.get('provincia', ''),       # <-- ESTO FALTABA
                'distrito': persona_local.get('distrito', ''),         # <-- ESTO FALTABA
                'mensaje': 'Persona detectada en el sistema local.'
            })

    # -------------------------------------------------------
    # FASE 2: BÚSQUEDA EXTERNA (RENIEC / SUNAT - API PÚBLICA)
    # -------------------------------------------------------
    try:
        if tipo_doc == 'DNI':
            url = f"https://api.apis.net.pe/v1/dni?numero={numero}"
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                data = response.json()
                return jsonify({
                    'success': True,
                    'origen': 'reniec',
                    'nombre': data.get('nombre', '')
                })
        elif tipo_doc == 'RUC':
            url = f"https://api.apis.net.pe/v1/ruc?numero={numero}"
            response = requests.get(url, timeout=5)
            if response.status_code == 200:
                data = response.json()
                return jsonify({
                    'success': True,
                    'origen': 'sunat',
                    'nombre': data.get('nombre', '')
                })
                
        return jsonify({'success': False, 'error': 'Documento no encontrado en registros oficiales externos.'})
    except Exception as e:
        return jsonify({'success': False, 'error': 'Error de conexión con el servicio externo.'})


# Diccionario temporal para guardar los códigos (En producción usarías Redis o la BD)
codigos_verificacion = {}

# ================== API DE VERIFICACIÓN DE CORREO ==================
@app.route('/api/enviar_codigo', methods=['POST'])
@roles_required(['admin', 'cajero'])
def enviar_codigo_correo():
    data = request.get_json()
    correo_destino = data.get('email')
    
    if not correo_destino:
        return jsonify({'error': 'Correo no proporcionado'}), 400

    # 1. Generamos un código de 4 dígitos
    codigo_otp = str(random.randint(1000, 9999))
    
    # Guardamos el código temporalmente asociado a ese correo
    codigos_verificacion[correo_destino] = codigo_otp

    # 2. Configuración de tu Gmail (Servidor Remitente)
    MI_CORREO = "arturo20010803@gmail.com" 
    MI_PASSWORD = "wxvqbwoflnoojhqh" 

    # 3. Armamos el mensaje
    mensaje = MIMEMultipart()
    mensaje['From'] = f"Taller Virgen de Chapi <{MI_CORREO}>"
    mensaje['To'] = correo_destino
    mensaje['Subject'] = f"Tu código de verificación: {codigo_otp}"

    cuerpo = f"""
    Hola,
    
    Para confirmar tu registro en el Taller Virgen de Chapi, 
    por favor bríndale este código de 4 dígitos al cajero:
    
    CÓDIGO: {codigo_otp}
    
    Si no solicitaste esto, puedes ignorar este mensaje.
    """
    mensaje.attach(MIMEText(cuerpo, 'plain'))

    try:
        # 4. Nos conectamos a Google y enviamos
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login(MI_CORREO, MI_PASSWORD)
        servidor.send_message(mensaje)
        servidor.quit()
        
        return jsonify({'success': True, 'mensaje': 'Código enviado exitosamente'})
    except Exception as e:
        print(f"Error enviando correo: {e}")
        return jsonify({'error': 'No se pudo enviar el correo. Verifique la dirección.'}), 500


@app.route('/api/verificar_codigo', methods=['POST'])
@roles_required(['admin', 'cajero'])
def verificar_codigo():
    data = request.get_json()
    correo = data.get('email')
    codigo_ingresado = data.get('codigo')
    
    # Comparamos el código que ingresó el cajero con el que guardamos
    if correo in codigos_verificacion and codigos_verificacion[correo] == codigo_ingresado:
        # Si es correcto, lo borramos por seguridad
        del codigos_verificacion[correo]
        return jsonify({'success': True})
    else:
        return jsonify({'error': 'Código incorrecto o expirado'}), 400

# ================== FUNCIÓN: AVISO DE VEHÍCULO LISTO ==================
def enviar_correo_vehiculo_listo(correo_cliente, nombre_cliente, placa, orden_id, informe_ia):
    if not correo_cliente or correo_cliente.strip() == "":
        return False # Si el cliente no dejó correo, cancelamos el envío en silencio

    MI_CORREO = "arturo20010803@gmail.com" 
    MI_PASSWORD = "wxvqbwoflnoojhqh"
    
    mensaje = MIMEMultipart()
    mensaje['From'] = f"Taller Virgen de Chapi <{MI_CORREO}>"
    mensaje['To'] = correo_cliente
    mensaje['Subject'] = f"✅ ¡Tu vehículo {placa} está listo! - Orden OT-{orden_id:04d}"

    # Diseñamos un correo corporativo en HTML
    cuerpo_html = f"""
    <html>
    <body style="font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f6f9; padding: 20px;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
            <div style="background-color: #2563eb; color: white; padding: 20px; text-align: center;">
                <h2 style="margin: 0; font-size: 24px;">Taller Virgen de Chapi</h2>
            </div>
            
            <div style="padding: 30px;">
                <h3 style="color: #1e293b; margin-top: 0;">¡Hola, {nombre_cliente}!</h3>
                <p style="color: #475569; font-size: 16px; line-height: 1.6;">
                    Te informamos que los servicios técnicos de tu vehículo con placa <strong style="color: #2563eb;">{placa}</strong> 
                    han sido concluidos exitosamente. Ya puedes pasar a recogerlo por nuestras instalaciones.
                </p>
                
                <div style="background-color: #f8fafc; border-left: 4px solid #10b981; padding: 15px; margin: 25px 0; border-radius: 0 8px 8px 0;">
                    <h4 style="margin: 0 0 10px 0; color: #10b981;">Resumen del Trabajo Realizado:</h4>
                    <p style="margin: 0; color: #334155; font-size: 14px; white-space: pre-wrap;">{informe_ia}</p>
                </div>
                
                <p style="color: #475569; font-size: 15px;">Por favor, acércate a nuestra área de caja indicando tu número de orden: <strong>OT-{orden_id:04d}</strong>.</p>
                
                <div style="margin-top: 30px; padding-top: 20px; border-top: 1px solid #e2e8f0; text-align: center;">
                    <p style="color: #94a3b8; font-size: 13px; margin: 0;">Gracias por confiar en nosotros.</p>
                </div>
            </div>
        </div>
    </body>
    </html>
    """
    mensaje.attach(MIMEText(cuerpo_html, 'html'))

    try:
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login(MI_CORREO, MI_PASSWORD)
        servidor.send_message(mensaje)
        servidor.quit()
        print(f"Correo enviado exitosamente a {correo_cliente}")
        return True
    except Exception as e:
        print(f"Error enviando correo de terminación: {e}")
        return False

# ==============================================================================
# FUNCIÓN DE CORREO: ALERTA DE EMERGENCIA A ADMINISTRADORES
# ==============================================================================
def enviar_correo_urgencia_admin(correo_admin, placa, orden_id, repuesto_faltante):
    # Usamos tus credenciales configuradas
    MI_CORREO = "arturo20010803@gmail.com" 
    MI_PASSWORD = "wxvqbwoflnoojhqh"
    
    mensaje = MIMEMultipart()
    mensaje['From'] = f"Alerta Taller Virgen de Chapi <{MI_CORREO}>"
    mensaje['To'] = correo_admin
    mensaje['Subject'] = f"🚨 URGENTE: Vehículo Pausado - OT-{orden_id:04d} (Placa: {placa})"

    # Diseño del correo en HTML (Corporativo pero de urgencia)
    cuerpo_html = f"""
    <html>
    <body style="font-family: 'Segoe UI', Arial, sans-serif; background-color: #f4f6f9; padding: 20px;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #ffffff; border: 2px solid #dc3545; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 15px rgba(220,53,69,0.2);">
            <div style="background-color: #dc3545; color: white; padding: 20px; text-align: center;">
                <h2 style="margin: 0; font-size: 24px;">🚨 ALERTA DE TALLER PAUSADO 🚨</h2>
            </div>
            
            <div style="padding: 30px;">
                <h3 style="color: #1e293b; margin-top: 0;">¡Atención Administración!</h3>
                <p style="color: #475569; font-size: 16px; line-height: 1.6;">
                    Un mecánico acaba de reportar la falta de un repuesto crítico. El trabajo en el vehículo ha sido <strong>detenido</strong>.
                </p>
                
                <div style="background-color: #fff3cd; border-left: 5px solid #ffc107; padding: 15px; margin: 25px 0; border-radius: 0 8px 8px 0;">
                    <h4 style="margin: 0 0 10px 0; color: #856404;">Detalles del Requerimiento:</h4>
                    <ul style="font-size: 15px; color: #333; margin: 0; padding-left: 20px;">
                        <li style="margin-bottom: 8px;"><strong>Nro de Orden:</strong> OT-{orden_id:04d}</li>
                        <li style="margin-bottom: 8px;"><strong>Placa del Vehículo:</strong> {placa}</li>
                        <li><strong>Requerimiento:</strong> <span style="color: #dc3545; font-weight: bold; font-size: 16px;">{repuesto_faltante}</span></li>
                    </ul>
                </div>
                
                <p style="color: #475569; font-size: 15px;">Por favor, gestione la compra de este insumo y marque la orden como "Comprado" en el Dashboard para que el mecánico pueda reanudar sus labores.</p>
            </div>
        </div>
    </body>
    </html>
    """
    mensaje.attach(MIMEText(cuerpo_html, 'html'))

    try:
        servidor = smtplib.SMTP('smtp.gmail.com', 587)
        servidor.starttls()
        servidor.login(MI_CORREO, MI_PASSWORD)
        servidor.send_message(mensaje)
        servidor.quit()
        print(f"Alerta enviada correctamente al admin: {correo_admin}")
    except Exception as e:
        print(f"Error enviando alerta a admin: {e}")

# ==============================================================================
# MÓDULO DE REPORTES GERENCIALES (DASHBOARD Y EXPORTACIÓN A EXCEL PREMIUM)
# ==============================================================================

@app.route('/reportes')
@roles_required(['admin'])
def reportes_index():
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    conn = get_db_connection()
    stats = {'ventas_mes': 0.0, 'ordenes_mes': 0, 'stock_critico': 0}
    ventas_lista = []
    mecanicos_lista = []
    dona_data = [0, 0, 0] # [Finalizados, En Proceso, Pausados]

    if conn:
        cursor = conn.cursor(dictionary=True)
        
        # --- 1. KPIs Generales ---
        cursor.execute("SELECT SUM(monto_total) as total FROM pagos WHERE MONTH(fecha_pago) = MONTH(CURDATE()) AND YEAR(fecha_pago) = YEAR(CURDATE()) AND estado != 'anulado'")
        res = cursor.fetchone()
        stats['ventas_mes'] = res['total'] if res['total'] else 0.0
        
        cursor.execute("SELECT COUNT(*) as total FROM ordenes_servicio WHERE MONTH(fecha_ingreso) = MONTH(CURDATE()) AND YEAR(fecha_ingreso) = YEAR(CURDATE())")
        stats['ordenes_mes'] = cursor.fetchone()['total']
        
        cursor.execute("SELECT COUNT(*) as total FROM productos WHERE tipo = 'repuesto' AND stock_actual <= stock_minimo")
        stats['stock_critico'] = cursor.fetchone()['total']
        
        # --- 2. Datos reales para el Gráfico de Dona ---
        cursor.execute("SELECT estado, COUNT(*) as cant FROM ordenes_servicio GROUP BY estado")
        for row in cursor.fetchall():
            if row['estado'] in ['finalizado', 'entregado']: dona_data[0] += row['cant']
            elif row['estado'] in ['en_proceso', 'pendiente']: dona_data[1] += row['cant']
            elif row['estado'] == 'esperando_repuesto': dona_data[2] += row['cant']

        # --- 3. Datos para Tabla Interactiva: Libro de Ingresos ---
        query_ventas = """
            SELECT pg.fecha_pago, pg.comprobante, p.numero_documento, p.nombre_completo as cliente, pg.metodo_pago, pg.monto_total, pg.estado
            FROM pagos pg
            JOIN ordenes_servicio os ON pg.orden_id = os.id
            JOIN vehiculos v ON os.vehiculo_id = v.id
            JOIN clientes c ON v.cliente_id = c.id
            JOIN personas p ON c.persona_id = p.id
            WHERE 1=1
        """
        params = []
        if fecha_inicio and fecha_fin:
            query_ventas += " AND DATE(pg.fecha_pago) BETWEEN %s AND %s"
            params.extend([fecha_inicio, fecha_fin])
        else:
            query_ventas += " AND pg.fecha_pago >= DATE_SUB(CURDATE(), INTERVAL 30 DAY)"
            
        query_ventas += " ORDER BY pg.fecha_pago DESC"
        cursor.execute(query_ventas, tuple(params))
        ventas_lista = cursor.fetchall()

        # --- 4. Datos para Tabla Interactiva: RRHH ---
        cursor.execute("""
            SELECT 
                p.nombre_completo as mecanico,
                COUNT(os.id) as total_asignadas,
                SUM(CASE WHEN os.estado IN ('entregado', 'finalizado') THEN 1 ELSE 0 END) as ordenes_completadas,
                SUM(pg.monto_total) as ingresos_generados
            FROM usuarios u
            JOIN personas p ON u.persona_id = p.id
            LEFT JOIN ordenes_servicio os ON os.usuario_id = u.id
            LEFT JOIN pagos pg ON pg.orden_id = os.id AND pg.estado != 'anulado'
            WHERE u.rol = 'mecanico' AND u.activo = 1
            GROUP BY u.id
            ORDER BY ingresos_generados DESC
        """)
        mecanicos_lista = cursor.fetchall()
        conn.close()

    return render_template('reportes/reportes.html', 
                           stats=stats, 
                           ventas=ventas_lista, 
                           mecanicos=mecanicos_lista,
                           dona_data=dona_data,
                           fecha_inicio=fecha_inicio, 
                           fecha_fin=fecha_fin)

@app.route('/reportes/exportar/ventas')
@roles_required(['admin'])
def exportar_ventas():
    # 1. Recibir las fechas del formulario HTML
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # 2. Consulta SQL Dinámica (Si hay fechas, filtra. Si no, trae todo)
    query = """
        SELECT pg.fecha_pago, pg.comprobante, p.numero_documento, p.nombre_completo as cliente, pg.metodo_pago, pg.monto_total, pg.estado
        FROM pagos pg
        JOIN ordenes_servicio os ON pg.orden_id = os.id
        JOIN vehiculos v ON os.vehiculo_id = v.id
        JOIN clientes c ON v.cliente_id = c.id
        JOIN personas p ON c.persona_id = p.id
        WHERE 1=1
    """
    params = []
    
    if fecha_inicio and fecha_fin:
        query += " AND DATE(pg.fecha_pago) BETWEEN %s AND %s"
        params.extend([fecha_inicio, fecha_fin])
        
    query += " ORDER BY pg.fecha_pago DESC"
    
    cursor.execute(query, tuple(params))
    ventas = cursor.fetchall()
    conn.close()

    # 3. Crear Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial de Ingresos"

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin', color='D0D7DE'), right=Side(style='thin', color='D0D7DE'), top=Side(style='thin', color='D0D7DE'), bottom=Side(style='thin', color='D0D7DE'))
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    headers = ['FECHA DE PAGO', 'NRO. COMPROBANTE', 'DNI / RUC', 'CLIENTE O RAZÓN SOCIAL', 'MÉTODO DE PAGO', 'TOTAL COBRADO', 'ESTADO']
    ws.append(headers)

    ws.row_dimensions[1].height = 25
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, v in enumerate(ventas, start=2):
        ws.cell(row=row_idx, column=1, value=v['fecha_pago'].strftime('%d/%m/%Y %I:%M %p') if v['fecha_pago'] else '---')
        ws.cell(row=row_idx, column=2, value=v['comprobante'])
        ws.cell(row=row_idx, column=3, value=v['numero_documento'])
        ws.cell(row=row_idx, column=4, value=v['cliente'])
        ws.cell(row=row_idx, column=5, value=v['metodo_pago'].upper())
        
        celda_monto = ws.cell(row=row_idx, column=6, value=float(v['monto_total']))
        celda_monto.number_format = '"S/" #,##0.00'
        
        estado = v['estado'].upper()
        celda_estado = ws.cell(row=row_idx, column=7, value=estado)
        if estado == 'ANULADO':
            celda_estado.font = Font(color="DC3545", bold=True)
        else:
            celda_estado.font = Font(color="198754", bold=True)

        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if row_idx % 2 == 0: cell.fill = zebra_fill
            if col_idx in [1, 2, 3, 5, 7]: cell.alignment = center_align
            else: cell.alignment = left_align

    anchos = {'A': 22, 'B': 20, 'C': 15, 'D': 40, 'E': 18, 'F': 18, 'G': 15}
    for col, ancho in anchos.items(): ws.column_dimensions[col].width = ancho
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    res = make_response(output.read())
    
    # Nombre de archivo dinámico si hay fechas
    if fecha_inicio and fecha_fin:
        nombre_archivo = f"Reporte_Ventas_{fecha_inicio}_al_{fecha_fin}.xlsx"
    else:
        nombre_archivo = "Reporte_Ingresos_Historico.xlsx"
        
    res.headers["Content-Disposition"] = f"attachment; filename={nombre_archivo}"
    res.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return res


@app.route('/reportes/exportar/inventario')
@roles_required(['admin'])
def exportar_inventario():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT codigo, nombre, categoria, marca, precio_compra, precio_venta, stock_actual, stock_minimo
        FROM productos WHERE tipo = 'repuesto'
        ORDER BY stock_actual ASC
    """)
    productos = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Estado de Inventario"

    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin', color='D0D7DE'), right=Side(style='thin', color='D0D7DE'), top=Side(style='thin', color='D0D7DE'), bottom=Side(style='thin', color='D0D7DE'))
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    alerta_roja_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    alerta_roja_font = Font(color="9C0006", bold=True)

    headers = ['CÓDIGO SISTEMA', 'DESCRIPCIÓN DEL REPUESTO', 'CATEGORÍA', 'MARCA', 'P. COMPRA', 'P. VENTA', 'STOCK', 'MÍNIMO', 'ESTADO']
    ws.append(headers)

    ws.row_dimensions[1].height = 25
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, p in enumerate(productos, start=2):
        ws.cell(row=row_idx, column=1, value=p['codigo'])
        ws.cell(row=row_idx, column=2, value=p['nombre'])
        ws.cell(row=row_idx, column=3, value=p['categoria'] if p['categoria'] else '---')
        ws.cell(row=row_idx, column=4, value=p['marca'] if p['marca'] else '---')
        
        celda_compra = ws.cell(row=row_idx, column=5, value=float(p['precio_compra']))
        celda_compra.number_format = '"S/" #,##0.00'
        
        celda_venta = ws.cell(row=row_idx, column=6, value=float(p['precio_venta']))
        celda_venta.number_format = '"S/" #,##0.00'
        
        celda_stock = ws.cell(row=row_idx, column=7, value=p['stock_actual'])
        ws.cell(row=row_idx, column=8, value=p['stock_minimo'])
        
        estado = "CRÍTICO" if p['stock_actual'] <= p['stock_minimo'] else "NORMAL"
        celda_estado = ws.cell(row=row_idx, column=9, value=estado)
        
        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if row_idx % 2 == 0: cell.fill = zebra_fill
            if col_idx in [1, 3, 4, 7, 8, 9]: cell.alignment = center_align
            else: cell.alignment = left_align

        if estado == "CRÍTICO":
            celda_estado.fill = alerta_roja_fill
            celda_estado.font = alerta_roja_font
            celda_stock.font = alerta_roja_font

    anchos = {'A': 18, 'B': 45, 'C': 20, 'D': 18, 'E': 14, 'F': 14, 'G': 10, 'H': 10, 'I': 14}
    for col, ancho in anchos.items(): ws.column_dimensions[col].width = ancho
    ws.auto_filter.ref = ws.dimensions

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    res = make_response(output.read())
    res.headers["Content-Disposition"] = "attachment; filename=Reporte_Inventario_VirgenDeChapi.xlsx"
    res.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return res


@app.route('/reportes/exportar/mecanicos')
@roles_required(['admin'])
def exportar_mecanicos():
    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)
    
    # Consulta corregida: Se cambió os.mecanico_id por os.usuario_id
    cursor.execute("""
        SELECT 
            p.nombre_completo as mecanico,
            p.numero_documento as dni,
            COUNT(os.id) as total_asignadas,
            SUM(CASE WHEN os.estado = 'entregado' OR os.estado = 'finalizado' THEN 1 ELSE 0 END) as ordenes_completadas,
            SUM(pg.monto_total) as ingresos_generados
        FROM usuarios u
        JOIN personas p ON u.persona_id = p.id
        LEFT JOIN ordenes_servicio os ON os.usuario_id = u.id
        LEFT JOIN pagos pg ON pg.orden_id = os.id AND pg.estado = 'valido'
        WHERE u.rol = 'mecanico'
        GROUP BY u.id
        ORDER BY ingresos_generados DESC
    """)
    mecanicos = cursor.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Rendimiento de Personal"

    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid") # Morado Corporativo RRHH
    header_font = Font(color="FFFFFF", bold=True, size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(left=Side(style='thin', color='D0D7DE'), right=Side(style='thin', color='D0D7DE'), top=Side(style='thin', color='D0D7DE'), bottom=Side(style='thin', color='D0D7DE'))
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

    headers = ['MECÁNICO / TÉCNICO', 'DNI', 'ÓRDENES ASIGNADAS', 'ÓRDENES COMPLETADAS', 'TASA EFECTIVIDAD', 'INGRESOS GENERADOS']
    ws.append(headers)

    ws.row_dimensions[1].height = 25
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row_idx, m in enumerate(mecanicos, start=2):
        ws.cell(row=row_idx, column=1, value=m['mecanico'])
        ws.cell(row=row_idx, column=2, value=m['dni'])
        ws.cell(row=row_idx, column=3, value=m['total_asignadas'])
        ws.cell(row=row_idx, column=4, value=m['ordenes_completadas'])
        
        # Calcular % de efectividad
        if m['total_asignadas'] > 0:
            efectividad = (m['ordenes_completadas'] / m['total_asignadas']) * 100
        else:
            efectividad = 0
            
        celda_efectividad = ws.cell(row=row_idx, column=5, value=efectividad/100) # Formato % en Excel
        celda_efectividad.number_format = '0.00%'
        
        celda_ingresos = ws.cell(row=row_idx, column=6, value=float(m['ingresos_generados'] or 0))
        celda_ingresos.number_format = '"S/" #,##0.00'

        ws.row_dimensions[row_idx].height = 20
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if row_idx % 2 == 0: cell.fill = zebra_fill
            if col_idx in [2, 3, 4, 5]: cell.alignment = center_align
            else: cell.alignment = left_align

    anchos = {'A': 40, 'B': 15, 'C': 25, 'D': 25, 'E': 20, 'F': 25}
    for col, ancho in anchos.items(): ws.column_dimensions[col].width = ancho

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    res = make_response(output.read())
    res.headers["Content-Disposition"] = "attachment; filename=Reporte_Rendimiento_Mecanicos.xlsx"
    res.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return res

# ================== MÓDULO DE INCIDENCIAS (ACTUALIZADO CON CORREOS) ==================

@app.route('/ordenes/reportar_falta', methods=['POST'])
@roles_required(['admin', 'mecanico'])
def reportar_falta_repuesto():
    orden_id = request.form['orden_id']
    repuesto = request.form['repuesto_faltante'] # Aquí llega todo el texto armado por el JS
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor(dictionary=True) # Usamos dictionary para extraer datos fácil
        try:
            # 1. Actualizar la orden a "esperando_repuesto"
            cursor.execute("""
                UPDATE ordenes_servicio 
                SET estado = 'esperando_repuesto', 
                    repuesto_faltante = %s, 
                    fecha_pausa = NOW() 
                WHERE id = %s
            """, (repuesto, orden_id))
            
            # 2. Obtener la placa del carro para el correo
            cursor.execute("""
                SELECT v.placa 
                FROM ordenes_servicio os
                JOIN vehiculos v ON os.vehiculo_id = v.id
                WHERE os.id = %s
            """, (orden_id,))
            vehiculo_bd = cursor.fetchone()
            placa = vehiculo_bd['placa'] if vehiculo_bd else 'Desconocida'

            # 3. MÁGIA: Buscar a TODOS los administradores que tengan un correo registrado
            cursor.execute("""
                SELECT p.email 
                FROM usuarios u
                JOIN personas p ON u.persona_id = p.id
                WHERE u.rol = 'admin' AND u.activo = 1 AND p.email IS NOT NULL AND p.email != ''
            """)
            administradores = cursor.fetchall()
            
            conn.commit()
            
            # 4. Enviar el correo a cada administrador encontrado
            for admin in administradores:
                enviar_correo_urgencia_admin(admin['email'], placa, int(orden_id), repuesto)

            flash(f"Trabajo pausado: Se alertó a los administradores sobre la necesidad de comprar: '{repuesto}'.", "warning")
            
        except Error as e:
            flash(f"Error al reportar la incidencia: {e.msg}", "danger")
        finally:
            cursor.close()
            conn.close()
            
    # Redirigimos a detalle_orden para que el mecánico vea que ya se puso en pausa
    return redirect(url_for('detalle_orden', orden_id=orden_id))


# --- REANUDAR TRABAJO (CUANDO YA SE COMPRÓ EL REPUESTO) ---
@app.route('/ordenes/reanudar', methods=['POST'])
@roles_required(['admin'])
def reanudar_trabajo():
    orden_id = request.form['orden_id']
    
    conn = get_db_connection()
    if conn:
        cursor = conn.cursor()
        try:
            # Regresamos el vehículo a "en_proceso"
            cursor.execute("""
                UPDATE ordenes_servicio 
                SET estado = 'en_proceso',
                    repuesto_faltante = NULL 
                WHERE id = %s
            """, (orden_id,))
            conn.commit()
            flash(f"✅ ¡Repuesto entregado! La Orden #{orden_id} ha vuelto a estado 'En Proceso'.", "success")
        except Error as e:
            flash(f"Error al reanudar la orden: {e.msg}", "danger")
        finally:
            cursor.close()
            conn.close()
            
    return redirect(url_for('dashboard'))

# ================== INICIO DEL SERVIDOR ==================
if __name__ == '__main__':
    app.run(debug=True, port=5002)